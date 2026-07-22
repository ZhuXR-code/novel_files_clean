"""FastAPI 主应用"""
import os
import sys
import time
import uuid
import shutil
import threading
import re
import pymysql
from datetime import datetime
from typing import List, Optional
from contextlib import asynccontextmanager

from pydantic import BaseModel
from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi import Request
from sqlalchemy import create_engine, text, func
from sqlalchemy.orm import Session, sessionmaker

from backend.logger import logger
from backend.models import Base, ScanConfig, ScanResult, FileMetadata, ColumnConfig, FileGroup, ParseLog, KeywordReplaceRule, HelpDoc
from backend.scanner import scan_files
from backend.regex_parser import parse_file_names_regex_only, parse_file_summary_regex_only
from backend import pipeline as pipeline_manager
from backend.keyword_replace import seed_default_rules

# 扫描进度追踪
scanning_progress = {}

# 解析进度追踪
parse_tasks = {}
parse_tasks_lock = threading.Lock()
parse_cancel_events = {}
parse_cancel_lock = threading.Lock()

# 解析日志
parse_logs = {}
parse_logs_lock = threading.Lock()

# 任务数据保留时间（完成/错误/取消后保留 30 分钟，然后自动清理）
TASK_TTL_SECONDS = 30 * 60


def _cleanup_old_tasks():
    """清理过期任务数据，防止全局字典无限增长导致内存泄漏。
    
    完成（done/error/cancelled）超过 TTL 的任务及其日志会被移除。
    此函数应在任务结束时调用，也可考虑定时调度。
    """
    now = datetime.now()
    with parse_tasks_lock:
        stale_ids = []
        for tid, task in parse_tasks.items():
            if task.get('status') in ('done', 'error', 'cancelled'):
                ended = task.get('ended_at')
                if ended and (now - ended).total_seconds() > TASK_TTL_SECONDS:
                    stale_ids.append(tid)
        for tid in stale_ids:
            del parse_tasks[tid]
    with parse_cancel_lock:
        for tid in stale_ids:
            parse_cancel_events.pop(tid, None)
    with parse_logs_lock:
        for tid in stale_ids:
            parse_logs.pop(tid, None)
    if stale_ids:
        # 清理数据库中的旧日志
        try:
            db = SessionLocal()
            db.query(ParseLog).filter(ParseLog.task_id.in_(stale_ids)).delete(synchronize_session=False)
            db.commit()
            db.close()
        except Exception as e:
            logger.warning(f'清理数据库旧日志失败: {e}')
        logger.debug(f'已清理 {len(stale_ids)} 个过期任务数据: {stale_ids}')

def add_parse_log(task_id: str, message: str, level: str = '', db_session_factory=None):
    """添加解析日志，同时记录到服务端日志文件和数据库"""
    if not level:
        msg_lower = message.lower() if message else ''
        if '失败' in message or '错误' in message or 'error' in msg_lower or 'fail' in msg_lower:
            level = 'error'
        elif '跳过' in message or '正则' in message or 'skip' in msg_lower:
            level = 'warn'
        else:
            level = 'info'

    # 同步记录到服务端日志
    if level == 'error':
        logger.error(f'[解析][{task_id}] {message}')
    elif level == 'warn':
        logger.warning(f'[解析][{task_id}] {message}')
    else:
        logger.info(f'[解析][{task_id}] {message}')

    # 写入内存缓存（供前端实时轮询）
    with parse_logs_lock:
        if task_id not in parse_logs:
            parse_logs[task_id] = []
        parse_logs[task_id].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'msg': message,
            'level': level,
        })
        if len(parse_logs[task_id]) > 500:
            parse_logs[task_id] = parse_logs[task_id][-500:]

    # 持久化到数据库
    if db_session_factory:
        try:
            db = db_session_factory()
            db.add(ParseLog(task_id=task_id, message=message, level=level))
            db.commit()
            db.close()
        except Exception as e:
            logger.warning(f'解析日志写入数据库失败: {e}')

# ===================== 数据库配置（优先从环境变量 / .env 读取） =====================
import os as _os

# 加载项目根目录下的 .env（若存在），避免将数据库凭据硬编码进代码
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

DB_USER = _os.environ.get('DB_USER', 'root')
# 安全加固：数据库密码禁止硬编码默认值，必须从环境变量或 .env 提供
DB_PASS = _os.environ.get('DB_PASS')
DB_HOST = _os.environ.get('DB_HOST', 'localhost')
DB_PORT = _os.environ.get('DB_PORT', '3308')
DB_NAME = _os.environ.get('DB_NAME', 'file_scanner_noai')

DATABASE_URL = f'mysql+pymysql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4'

# 先连接MySQL（不指定数据库）来自动创建数据库
INIT_URL = f'mysql+pymysql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}?charset=utf8mb4'

# 本地 EXE 模式：使用内置 SQLite，不依赖外部 MySQL
from backend.db_config import IS_SQLITE, sqlite_db_path, app_data_root

engine = None
SessionLocal = None


def make_pymysql_conn():
    """创建 pymysql 原生连接，用于工程类高并发解析的批量 upsert 写入。

    说明：SQLAlchemy 的 engine URL 已使用 mysql+pymysql 驱动，
    此处为批量写入路径显式使用 pymysql 原生连接，以便拼接
    INSERT ... ON DUPLICATE KEY UPDATE 批量语句一次性写入。

    关键：设置 connect/read/write 超时，避免数据库锁等待或网络抖动时
    cur.execute 无限阻塞（此前表现为"扫一批后卡住"）。将 read/write 超时从
    120s 下调到 30s，并在批量写入处配套重试，避免长时间无进度。
    """
    logger.debug(f'创建 pymysql 连接: {DB_HOST}:{DB_PORT}/{DB_NAME}')
    t0 = time.perf_counter()
    try:
        conn = pymysql.connect(
            host=DB_HOST,
            port=int(DB_PORT),
            user=DB_USER,
            password=DB_PASS,
            database=DB_NAME,
            charset='utf8mb4',
            autocommit=False,
            connect_timeout=10,     # 建立连接最多 10s
            read_timeout=30,        # 读取/执行结果最多 30s（避免锁等待时长时间无进度）
            write_timeout=30,       # 发送数据最多 30s
        )
        logger.debug(f'[数据库] pymysql 连接创建成功, 耗时={time.perf_counter() - t0:.3f}s')
        return conn
    except Exception as e:
        logger.error(f'[数据库] 创建 pymysql 连接失败(耗时={time.perf_counter() - t0:.3f}s): {e}', exc_info=True)
        raise


def _is_path_under_root(root: str, path: str) -> bool:
    """判断 path 是否位于 root 目录（含子目录）内，防止路径穿越绕过。

    使用 os.path.commonpath 替代脆弱的 startswith 比较
    （原 startswith 可被 'c:\\docs_evil\\file.txt' 对 root='c:\\docs' 绕过）。
    """
    try:
        root_abs = os.path.abspath(root)
        path_abs = os.path.abspath(path)
        return os.path.commonpath([root_abs, path_abs]) == root_abs
    except ValueError:
        # 跨盘符等异常路径，直接拒绝
        return False


def init_database():
    """初始化数据库和表"""
    global engine, SessionLocal

    # 本地 EXE 模式：使用内置 SQLite，跳过全部 MySQL 专属逻辑（建库/迁移）
    if IS_SQLITE:
        _init_sqlite_database()
        return

    if not DB_PASS:
        raise RuntimeError(
            '数据库密码未配置：请通过环境变量 DB_PASS 或在项目根目录 .env 中设置 DB_PASS 后再启动服务。'
        )

    logger.info(f'正在连接MySQL: {DB_HOST}:{DB_PORT}')
    
    # 第一步：创建数据库（如果不存在）
    try:
        temp_engine = create_engine(INIT_URL, pool_pre_ping=True)
        with temp_engine.connect() as conn:
            conn.execute(text(f"CREATE DATABASE IF NOT EXISTS `{DB_NAME}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"))
            conn.commit()
        temp_engine.dispose()
        logger.info(f'数据库 [{DB_NAME}] 已就绪')
    except Exception as e:
        logger.error(f'创建数据库失败: {e}')
        raise

    # 第二步：连接数据库并创建表
    engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_size=10, max_overflow=20, pool_recycle=3600)
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    
    Base.metadata.create_all(bind=engine)
    logger.info('数据表创建完成')

    # 迁移：修复 FileMetadata 外键约束（确保 ON DELETE CASCADE）
    _migrate_fk_cascade()

    # 迁移：为 FileMetadata 表添加新列（如果不存在）
    _migrate_file_metadata()

    # 迁移：删除 file_metadata 中的 copywriting 列（已合并到 summary）
    _migrate_drop_copywriting()

    # 迁移：更新/新增所有表字段的中文注释。
    # 注意：注释迁移会用 ALTER ... MODIFY COLUMN 补注释，该语句未带
    # AUTO_INCREMENT，MySQL 重建表时可能丢失主键自增属性。故必须放在
    # 自增迁移之前执行，让自增迁移在最后兜底修复。
    _migrate_column_comments()

    # 迁移：修复 file_metadata 表 id 字段缺少 AUTO_INCREMENT 的问题
    _migrate_file_metadata_id_autoincrement()

    # 迁移：修复所有元数据表主键 id 缺少 AUTO_INCREMENT 的问题（避免 INSERT 报 1364）
    # 必须放在注释迁移之后，作为最后兜底。
    _migrate_all_id_autoincrement()

    # 迁移：添加索引优化查询性能（仅缺失时添加）
    _migrate_indexes()

    # 迁移：为 scan_configs 表添加自定义名称列（若已存在则跳过）
    _migrate_scan_config_name()

    # 迁移：为 scan_configs 表添加 parse_on_scan 列（扫描时同步工程类解析）
    _migrate_scan_config_parse_on_scan()

    # 第三步：初始化默认数据
    _init_default_data()


def _init_sqlite_database():
    """本地 EXE 模式：使用内置 SQLite 数据库（无需外部 MySQL）。

    直接按 models.py 的表结构建表（已含外键/唯一约束/索引），
    然后初始化默认数据（列配置、默认扫描配置、帮助文档）。
    不做任何 MySQL 专属的自动迁移。
    """
    global engine, SessionLocal
    import os as _os_mod
    db_path = sqlite_db_path()
    logger.info(f'正在初始化 SQLite 数据库: {db_path}')
    _os_mod.makedirs(_os_mod.path.dirname(db_path), exist_ok=True)

    url = f'sqlite:///{db_path}'
    engine = create_engine(
        url,
        connect_args={'check_same_thread': False},
        pool_pre_ping=True,
    )
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

    Base.metadata.create_all(bind=engine)
    logger.info('SQLite 数据表创建完成')

    # 迁移：为旧 SQLite 库补 parse_on_scan 列（create_all 不会给已存在的表加列）
    _migrate_scan_config_parse_on_scan()

    _init_default_data()
    logger.info(f'SQLite 初始化完成: {db_path}')


def _migrate_drop_copywriting():
    """迁移：删除 file_metadata.copywriting 列"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text(
                "SELECT COUNT(*) FROM information_schema.columns WHERE table_schema = :db AND table_name = 'file_metadata' AND column_name = 'copywriting'"
            ), {'db': DB_NAME})
            if result.scalar() > 0:
                conn.execute(text("ALTER TABLE file_metadata DROP COLUMN copywriting"))
                logger.info('迁移: 删除列 file_metadata.copywriting')
            conn.commit()
    except Exception as e:
        logger.warning(f'删除copywriting列警告: {e}')


def _migrate_file_metadata_id_autoincrement():
    """迁移：修复 file_metadata 表 id 字段缺少 AUTO_INCREMENT 的问题"""
    try:
        with engine.connect() as conn:
            # 检查 file_metadata 表是否存在
            result = conn.execute(text(
                "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = :db AND table_name = 'file_metadata'"
            ), {'db': DB_NAME})
            if result.scalar() == 0:
                return

            # 检查 id 字段是否有 auto_increment
            result = conn.execute(text(
                "SELECT EXTRA FROM information_schema.columns "
                "WHERE table_schema = :db AND table_name = 'file_metadata' AND column_name = 'id'"
            ), {'db': DB_NAME})
            row = result.fetchone()
            if row and 'auto_increment' in (row[0] or ''):
                return  # 已有 AUTO_INCREMENT，无需处理

            # 修复：添加 AUTO_INCREMENT
            conn.execute(text("ALTER TABLE file_metadata MODIFY COLUMN id INT NOT NULL AUTO_INCREMENT"))
            conn.commit()
            logger.info('迁移: 修复 file_metadata.id 添加 AUTO_INCREMENT')
    except Exception as e:
        logger.warning(f'file_metadata id auto_increment 迁移警告: {e}')


def _migrate_all_id_autoincrement():
    """迁移：确保所有元数据表的主键 id 列都是 AUTO_INCREMENT。

    历史数据库中存在部分表（column_configs、file_groups、scan_configs 等）
    主键缺少 AUTO_INCREMENT，导致启动时 INSERT 报
    (1364, "Field 'id' doesn't have a default value")。此处对所有表统一修复。
    """
    try:
        from sqlalchemy import Integer as _Integer
        with engine.connect().execution_options(isolation_level="AUTOCOMMIT") as conn:
            for table_name, table in Base.metadata.tables.items():
                pk_cols = [c for c in table.columns if c.primary_key]
                if len(pk_cols) != 1:
                    continue
                pk = pk_cols[0]
                if not isinstance(pk.type, _Integer):
                    continue
                # 表是否存在
                res = conn.execute(text(
                    "SELECT COUNT(*) FROM information_schema.tables "
                    "WHERE table_schema = :db AND table_name = :t"
                ), {'db': DB_NAME, 't': table_name})
                if res.scalar() == 0:
                    continue
                # 用 SHOW CREATE TABLE 判断 id 是否真为自增。
                # 注意：information_schema.EXTRA 存在"假象"——显示有 auto_increment
                # 但 ALTER 因列类型未变被 MySQL in-place 优化跳过，实际未生效，
                # 导致后续 INSERT 仍报 1364。故以建表语句为真实依据。
                res = conn.execute(text(f"SHOW CREATE TABLE `{table_name}`"))
                create_sql = res.fetchone()[1]
                if 'AUTO_INCREMENT' in create_sql:
                    continue
                # 修复：必须加 ALGORITHM=COPY 强制重建表，否则 MySQL 会
                # in-place 跳过，自增属性不生效（INSERT 仍报 1364）。
                conn.execute(text(
                    f"ALTER TABLE `{table_name}` MODIFY COLUMN `{pk.name}` "
                    f"INT NOT NULL AUTO_INCREMENT, ALGORITHM=COPY"
                ))
                logger.info(f'迁移: 修复 {table_name}.{pk.name} 添加 AUTO_INCREMENT')
    except Exception as e:
        logger.warning(f'全表 id auto_increment 迁移警告: {e}')


def _migrate_column_comments():
    """迁移：仅对缺失注释的表字段补充中文注释（先检查，避免重复ALTER）"""
    try:
        with engine.connect() as conn:
            # 检查表是否存在
            tables_result = conn.execute(text(
                "SELECT TABLE_NAME FROM information_schema.tables WHERE table_schema = :db"
            ), {'db': DB_NAME})
            existing_tables = {row[0] for row in tables_result}

            # 按表定义字段注释（只包含缺失的）
            table_columns = {
                'scan_configs': [
                    ('id', 'INT', '扫描配置ID'),
                    ('folder_path', 'VARCHAR(500)', '扫描文件夹路径'),
                    ('file_types', 'VARCHAR(200)', '文件类型，多个用逗号分隔'),
                    ('excluded_folders', 'TEXT', '排除的文件夹，多个用逗号分隔'),
                    ('created_at', 'DATETIME', '创建时间'),
                    ('updated_at', 'DATETIME', '更新时间'),
                ],
                'scan_results': [
                    ('id', 'INT', '扫描结果ID'),
                    ('file_name', 'VARCHAR(500)', '完整文件名（含后缀）'),
                    ('file_size', 'BIGINT', '文件大小（字节）'),
                    ('file_path', 'VARCHAR(1000)', '文件完整路径'),
                    ('created_date', 'DATETIME', '文件创建日期'),
                    ('scan_config_id', 'INT', '关联扫描配置ID'),
                    ('scanned_at', 'DATETIME', '扫描时间'),
                ],
                'file_metadata': [
                    ('id', 'INT', '元数据ID'),
                    ('scan_result_id', 'INT', '关联扫描结果ID'),
                    ('novel_name', 'VARCHAR(500)', '小说名'),
                    ('author', 'VARCHAR(200)', '作者'),
                    ('summary', 'TEXT', '内容简介/摘要'),
                    ('parsed_at', 'DATETIME', '解析时间'),
                ],
                'column_configs': [
                    ('id', 'INT', '列配置ID'),
                    ('column_key', 'VARCHAR(50)', '列标识'),
                    ('display_name', 'VARCHAR(100)', '列显示名称'),
                    ('visible', 'TINYINT(1)', '是否显示'),
                    ('sort_order', 'INT', '排序顺序'),
                ],
            }

            altered_any = False
            for table_name, columns in table_columns.items():
                if table_name not in existing_tables:
                    continue
                # 先查该表已有字段注释
                existing_result = conn.execute(text(
                    "SELECT COLUMN_NAME, COLUMN_COMMENT FROM information_schema.columns "
                    "WHERE table_schema = :db AND table_name = :table"
                ), {'db': DB_NAME, 'table': table_name})
                existing_comments = {row[0]: row[1] for row in existing_result}

                for col_name, col_type, col_comment in columns:
                    if existing_comments.get(col_name) == col_comment:
                        continue  # 注释已存在，跳过
                    try:
                        conn.execute(text(
                            f"ALTER TABLE `{table_name}` MODIFY COLUMN `{col_name}` {col_type} COMMENT :comment"
                        ), {'comment': col_comment})
                        altered_any = True
                    except Exception:
                        pass  # 某些字段可能因类型不匹配跳过
            if altered_any:
                conn.commit()
                logger.info('迁移: 部分字段已补充中文注释')
            else:
                logger.info('迁移: 所有字段注释已完整，无需修改')
    except Exception as e:
        logger.warning(f'列注释迁移警告: {e}')


def _migrate_fk_cascade():
    """迁移FileMetadata表的外键约束，确保ON DELETE CASCADE"""
    try:
        with engine.connect() as conn:
            # 检查 file_metadata 表是否存在
            result = conn.execute(text(
                "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = :db AND table_name = 'file_metadata'"
            ), {'db': DB_NAME})
            if result.scalar() == 0:
                return

            # 检查现有外键的 DELETE_RULE
            result = conn.execute(text(
                "SELECT CONSTRAINT_NAME, DELETE_RULE FROM information_schema.REFERENTIAL_CONSTRAINTS "
                "WHERE constraint_schema = :db AND table_name = 'file_metadata' "
                "AND referenced_table_name = 'scan_results'"
            ), {'db': DB_NAME})
            row = result.fetchone()
            if row and row[1] == 'CASCADE':
                return  # 已经是 CASCADE，无需处理

            if row:
                # 删除旧外键，重新添加带 CASCADE 的
                fk_name = row[0]
                conn.execute(text(f"ALTER TABLE file_metadata DROP FOREIGN KEY `{fk_name}`"))
                conn.execute(text(
                    "ALTER TABLE file_metadata ADD CONSTRAINT `{0}` "
                    "FOREIGN KEY (scan_result_id) REFERENCES scan_results(id) ON DELETE CASCADE".format(fk_name)
                ))
                conn.commit()
                logger.info(f'迁移: 修复外键 {fk_name} -> ON DELETE CASCADE')
    except Exception as e:
        logger.warning(f'外键迁移警告: {e}')


def _migrate_file_metadata():
    """迁移FileMetadata表，添加新增的列"""
    new_columns = [
        'summary TEXT',
        'progress VARCHAR(200)',
        'source VARCHAR(100)',
    ]
    try:
        with engine.connect() as conn:
            # 检查表是否存在
            result = conn.execute(text(
                "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = :db AND table_name = 'file_metadata'"
            ), {'db': DB_NAME})
            if result.scalar() == 0:
                return

            # 获取现有列
            result = conn.execute(text(
                "SELECT COLUMN_NAME FROM information_schema.columns WHERE table_schema = :db AND table_name = 'file_metadata'"
            ), {'db': DB_NAME})
            existing_cols = {row[0] for row in result}

            # 添加缺少的列
            for col_def in new_columns:
                col_name = col_def.split()[0]
                if col_name not in existing_cols:
                    conn.execute(text(f"ALTER TABLE file_metadata ADD COLUMN {col_def}"))
                    logger.info(f'迁移: 添加列 file_metadata.{col_name}')
            conn.commit()
    except Exception as e:
        logger.warning(f'数据库迁移警告: {e}')


def _migrate_indexes():
    """迁移：为所有表添加缺失的索引"""
    indexes = [
        # scan_results
        ('scan_results', 'idx_scan_config_id', ['scan_config_id']),
        ('scan_results', 'idx_file_name', ['file_name(255)']),
        ('scan_results', 'idx_file_path', ['file_path(255)']),
        ('scan_results', 'idx_config_id_file_size', ['scan_config_id', 'file_size']),
        ('scan_results', 'idx_created_date', ['created_date']),
        ('scan_results', 'idx_scanned_at', ['scanned_at']),
        # file_metadata
        ('file_metadata', 'idx_novel_name', ['novel_name(255)']),
        ('file_metadata', 'idx_author', ['author']),
        # scan_configs
        ('scan_configs', 'idx_config_folder_path', ['folder_path(255)']),
        ('scan_configs', 'idx_config_created_at', ['created_at']),
        # column_configs
        ('column_configs', 'idx_column_sort_order', ['sort_order']),
        # file_groups
        ('file_groups', 'idx_group_config_novel', ['config_id', 'novel_name']),
        ('file_groups', 'idx_group_config_count', ['config_id', 'file_count']),
    ]
    tables = ['scan_results', 'file_metadata', 'scan_configs', 'column_configs', 'file_groups']
    try:
        with engine.connect() as conn:
            result = conn.execute(text(
                "SELECT DISTINCT INDEX_NAME FROM information_schema.statistics "
                "WHERE table_schema = :db AND table_name IN :tables"
            ), {'db': DB_NAME, 'tables': tuple(tables)})
            existing_idx = {row[0] for row in result}
            for table, idx_name, cols in indexes:
                if idx_name not in existing_idx:
                    col_def = ', '.join(cols)
                    conn.execute(text(f"CREATE INDEX {idx_name} ON {table}({col_def})"))
                    logger.info(f'迁移: 创建索引 {table}.{idx_name}')
            conn.commit()
    except Exception as e:
        logger.warning(f'索引迁移警告: {e}')


def _migrate_scan_config_name():
    """迁移：为 scan_configs 表添加自定义名称列 name（若已存在则跳过）"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text(
                "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = :db AND table_name = 'scan_configs'"
            ), {'db': DB_NAME})
            if result.scalar() == 0:
                return
            result = conn.execute(text(
                "SELECT COLUMN_NAME FROM information_schema.columns WHERE table_schema = :db AND table_name = 'scan_configs'"
            ), {'db': DB_NAME})
            existing_cols = {row[0] for row in result}
            if 'name' not in existing_cols:
                conn.execute(text("ALTER TABLE scan_configs ADD COLUMN name VARCHAR(200) NULL DEFAULT '' COMMENT '扫描配置自定义名称'"))
                logger.info('迁移: 添加列 scan_configs.name')
            conn.commit()
    except Exception as e:
        logger.warning(f'scan_configs name 迁移警告: {e}')


def _migrate_scan_config_parse_on_scan():
    """迁移：为 scan_configs 表添加 parse_on_scan 列（扫描时同步工程类解析）。

    旧库升级时补列，避免已有 SQLite/MySQL 库因缺列导致查询崩溃。
    使用 SQLAlchemy inspector，兼容 MySQL 与 SQLite。
    """
    try:
        from sqlalchemy import inspect as sa_inspect
        insp = sa_inspect(engine)
        if 'scan_configs' not in insp.get_table_names():
            return
        existing_cols = {c['name'] for c in insp.get_columns('scan_configs')}
        if 'parse_on_scan' in existing_cols:
            return
        with engine.begin() as conn:
            if IS_SQLITE:
                conn.execute(text("ALTER TABLE scan_configs ADD COLUMN parse_on_scan BOOLEAN NOT NULL DEFAULT 1"))
            else:
                conn.execute(text("ALTER TABLE scan_configs ADD COLUMN parse_on_scan TINYINT(1) NOT NULL DEFAULT 1 COMMENT '扫描时是否同步执行工程类解析'"))
        logger.info('迁移: 添加列 scan_configs.parse_on_scan')
    except Exception as e:
        logger.warning(f'scan_configs parse_on_scan 迁移警告: {e}')


def _init_default_data():
    """初始化默认配置数据"""
    db = SessionLocal()
    try:
        # 初始化列配置（新增或更新）
        default_columns = [
            {'column_key': 'file_name', 'display_name': '文件名', 'visible': True, 'sort_order': 1},
            {'column_key': 'file_size', 'display_name': '文件大小', 'visible': True, 'sort_order': 2},
            {'column_key': 'file_path', 'display_name': '文件路径', 'visible': True, 'sort_order': 3},
            {'column_key': 'created_date', 'display_name': '创建日期', 'visible': True, 'sort_order': 4},
            {'column_key': 'novel_name', 'display_name': '小说名', 'visible': True, 'sort_order': 5},
            {'column_key': 'author', 'display_name': '作者', 'visible': True, 'sort_order': 6},
            {'column_key': 'summary', 'display_name': '内容简介', 'visible': False, 'sort_order': 7},
            {'column_key': 'progress', 'display_name': '进度', 'visible': True, 'sort_order': 8},
            {'column_key': 'source', 'display_name': '来源', 'visible': True, 'sort_order': 9},
    ]
        existing_cols = {c.column_key: c for c in db.query(ColumnConfig).all()}
        for col in default_columns:
            if col['column_key'] not in existing_cols:
                db.add(ColumnConfig(**col))
                logger.info(f'新增列配置: {col["display_name"]}')
        db.commit()

        # 如果没有扫描配置，创建一个默认的
        if db.query(ScanConfig).count() == 0:
            default_path = os.path.expanduser('~/Documents')
            db.add(ScanConfig(
                folder_path=default_path,
                file_types='txt',
                excluded_folders='',
            ))
            db.commit()
            logger.info(f'默认扫描配置已初始化: {default_path}')

        # 同步帮助手册默认文档（以代码内置内容为准，已存在的按 doc_key 覆盖更新）
        _seed_help_docs(db)

        # 补齐缺失的预置关键词替换规则（去水印等），幂等，不影响用户自定义规则
        seed_default_rules(db)

        # 启动时重建所有已有配置的文件分组
        config_ids = [c.id for c in db.query(ScanConfig.id).all()]
        for cid in config_ids:
            rebuild_file_groups(cid, db)

    finally:
        db.close()


def _seed_help_docs(db: Session):
    """写入帮助手册默认文档（Markdown 内容）"""
    docs = [
        {
            'doc_key': 'quickstart',
            'title': '快速上手（一步步操作指南）',
            'sort_order': 1,
            'content': '''# 快速上手：一步步操作指南

本系统用于把本地小说 `TXT / MD` 文件 **扫描入库 → 解析书名作者 → 整理去重 → 导出 / 清理**。
所有解析都是**本地正则方法**，**无需联网、无需配置任何密钥、不产生任何费用**。

下面按推荐顺序说明每一步的 **操作 / 效果 / 注意事项**。

---

## 第 1 步：创建扫描配置
- **操作**：左侧「扫描配置」区点「+ 新增配置」，填写：
  - **配置名称**：自定义，便于识别（如「待整理-番外合集」）
  - **文件夹路径**：要扫描的本地目录（如 `D:\\小说\\待整理`）
  - **文件类型**：勾选 `txt` / `md`，或自定义扩展名，**至少选一个**
  - **排除文件夹**（可选）：逗号分隔，这些子目录不会被扫描（如 `已整理,备份`）
  填完点「保存」。
- **效果**：配置出现在左侧列表，成为后续所有操作的作用对象。
- **注意**：
  - 文件夹路径必须**真实存在且有读取权限**，否则扫描会失败。
  - 排除文件夹可避免把「已整理」「备份」等目录重复入库。
  - 一个配置对应**一个目录**；多个目录请建多个配置。

## 第 2 步：扫描（把文件读入数据库）
- **操作**：在左侧对应配置上点「扫描」按钮。
- **效果**：系统遍历目录，把符合条件的文件作为记录写入数据库，进度条实时显示已处理数量；扫描结束会自动重建「合集」。
- **注意**：
  - 扫描是**增量**的：已存在的文件不会重复入库，只新增新文件。
  - 扫描只登记文件名 / 路径 / 大小等元信息，**不会解析书名作者，也不会改动你的源文件**。
  - 大目录扫描可能较久，请等进度条走完再继续。
  - 扫描的具体执行步骤、增量去重、排除目录匹配与"扫描阶段"关键词替换的生效时机，详见《扫描逻辑详解》。

## 第 3 步：工程解析（提取书名 / 作者 / 进度 / 来源）
- **操作**：在「数据列表」页用工具栏的「工程解析」选择范围，依次点：
  1. **文件名解析**
  2. **摘要提取**
  - 范围三选一：
    - **已选**：只对勾选的行解析
    - **全部**：对本配置下所有文件解析（已有结果会跳过）
    - **强制重跑全部**：忽略已有结果，全部重新解析
- **效果**：
  - **文件名解析**：从文件名用正则提取 `书名 / 作者 / 进度 / 来源` 四列。
  - **摘要提取**：读取每个文件正文开头，截取首章之前的简介填入「摘要」列。
- **注意**：
  - 文件名解析**只依赖文件名格式**，提取效果取决于原文件名是否规范（如「书名 作者.txt」「《书名》作者.txt」）。
  - 摘要提取会**读取文件内容**，文件多 / 大时较慢，但仍是本地操作。
  - 个别文件解析失败不会中断任务，只是对应列为空，可稍后用「强制重跑全部」重试。

## 第 4 步：查看、筛选与编辑
- **操作**：
  - 顶部搜索框按 `文件名 / 路径 / 小说名 / 作者` 模糊搜索；
  - 工具栏「小说筛选」可按书名筛选；
  - 点「列显示配置」勾选要展示的字段；
  - 行内可**双击编辑** `书名 / 作者 / 摘要 / 进度 / 来源`。
- **效果**：快速定位目标文件，并手工修正自动解析不准的字段。
- **注意**：手工编辑会覆盖自动解析结果，且**不会被后续普通解析覆盖**（除非用「强制重跑全部」）。

## 第 5 步：导出
- **操作**：
  - 勾选若干行 → 点「导出 MD」：把选中文件解析结果导出为 Markdown（每个文件一个 `.md`，存到 `exports/` 目录）。
  - 点「导出 Excel」→ 选「选中」或「全部」：导出为 Excel 表格。
- **效果**：生成可供查阅 / 归档的文件，列含 `文件名 / 小说名 / 作者 / 进度 / 来源 / 摘要`。
- **注意**：导出的是数据库里的解析结果，**不影响源文件**。

## 第 6 步（可选）：合集与标记重复
- **操作**：在「数据列表」页点「合集模式」按钮，文件按小说名聚合为合集；点「标记重复」一键勾选应清理的冗余行。
- **效果**：在「同一合集内、按 (作者+小说名) 子分组」套用五则规则智能勾选——① 五字段完全相等的多本删旧留新；② 进度全为纯数字则留进度最大者；③ 含中文进度(如完结)不删，但若存在更小的"进度最大文件"且组内有完结版则连它一起删；④ 本组内唯一最大文件始终保留；⑤ 进度为「完结+数字番外」组合时按数字排序，数字最大者不删、其余删（被删者若恰为本组文件大小最大者也不删）。详细规则与举例见《功能说明》第 4 节。
- **注意**：标记只是「勾选」，**并不删除**；真正删除要走删除或一键清理（见下）。

## 第 7 步（可选）：设置关键词替换
- **操作**：切到「设置」页，分别维护「扫描阶段」和「解析阶段」的替换规则（`查找 → 替换为`，支持多条按顺序排列）。
- **效果**：
  - **扫描阶段**：在**入库前**改写文件名（如去掉 `[XX论坛]` 前缀）。
  - **解析阶段**：在**解析后**规范化 `书名 / 作者 / 进度 / 来源`（如把「更78」规整为「更」）。
- **注意**：规则按 `sort_order` 顺序生效；改完规则后，建议对相关配置「强制重跑」对应解析以套用新规则。

> 想一步到位？见《功能说明》里的 **「一键清理」**：一次完成 `扫描 → 解析 → 生成合集 → 标记重复 → 删除确认` 全流程。''',
        },
        {
            'doc_key': 'scan_logic',
            'title': '扫描逻辑详解',
            'sort_order': 2,
            'content': '''# 扫描逻辑详解

本页说明系统「扫描」时到底做了什么，分**普通操作（手动扫描）**与**一键操作（一键清理的扫描节点）**两种触发方式，二者底层调用同一套扫描函数（`backend/scanner.py` 的 `scan_files`）。

---

## 一、两种触发方式与效果

### 普通操作：手动扫描
- **操作**：在「扫描配置」列表里，对应配置上点「扫描」按钮（调用 `/api/scan/{config_id}`）。
- **效果**：遍历该配置的文件夹，把符合条件的文件作为记录写入数据库，进度条实时显示已处理数量；扫描结束会自动重建「合集」。
- **注意事项**：
  - 扫描是**增量**的：已存在的文件不会重复入库，只新增新文件。
  - 扫描只登记文件名 / 路径 / 大小 / 创建时间等元信息，**不会解析书名作者，也不会改动你的源文件**。
  - 大目录扫描可能较久，请等进度条走完再继续，避免重复点击。

### 一键操作：一键清理的扫描节点（节点 1）
- **效果**：一键清理流程串行自动执行，扫描作为第 1 个节点先跑，跑完紧接着自动做「工程解析 → 生成合集 → 标记重复 → 二次确认删除」。
- **注意事项**：
  - 扫描逻辑与普通操作完全一致（同一函数），所以增量、不改源文件等特性同样适用。
  - 若一键清理被取消，已完成的扫描结果**保留、不回滚**，后续节点跳过。
  - 删除模式（`db` / `file`）在启动一键清理时选定，扫描阶段不会删除任何文件。

---

## 二、系统实际做了什么（逐步拆解）

不论哪种触发方式，扫描都会按下面顺序执行：

1. **校验目录**：检查配置里的 `文件夹路径` 是否真实存在；不存在则记录错误并中止（手动扫描会报错提示，不会静默失败）。
2. **加载「扫描阶段」关键词替换规则**：从数据库读取 `scope=scan` 且已启用的规则，按 `sort_order` 升序排列。这些规则会在**写入数据库之前**用于改写文件名（如去掉 `[XX论坛]` 前缀）。
3. **加载去重基准**：读取该配置下已有的全部 `file_path`，构造"已存在路径集合"。**Windows 下路径不区分大小写**，统一转小写后再比对，用于跳过已入库的文件。
4. **规范化文件类型**：把配置里的 `文件类型`（如 `txt, md`）拆成集合，去空格、转小写、去掉前导点，得到最终匹配扩展名集合（如 `{txt, md}`）。
5. **构建排除集合**：把 `排除文件夹`（逗号分隔）拆成两类：
   - **纯文件夹名**：按名匹配任意层级的子目录（如 `已整理` 会排除所有层级的 `已整理` 目录）；
   - **相对路径**：若排除项是相对扫描根目录的路径，则按相对路径匹配。
6. **递归遍历目录树**：使用 `os.walk` 自上而下遍历，遇到排除项直接**剪枝**（不再进入该目录），并跳过中间路径被排除的目录，从而既不扫入也不浪费时间进入。
7. **逐文件处理（仅匹配扩展名的文件）**：
   - 若文件路径已命中"去重基准集合" → **跳过**，不重复入库；
   - 否则用 `os.stat` 读取**文件大小**与**创建时间**（读取失败时记录大小为 0、创建时间为空，不中断整体任务）；
   - 应用第 2 步的「扫描阶段」关键词替换规则，**改写写入数据库的 `file_name`**；
   - 生成一条 `ScanResult` 记录（含 `file_name / file_path / file_size / created_date / scan_config_id`）。
8. **批量入库**：每累计 **50 条**执行一次 `bulk_save_objects` + `commit`；遍历结束后把剩余批次再次提交，避免丢数据。
9. **实时进度**：每一步都会把已处理数量写入进度字典，前端进度条据此刷新显示。
10. **返回结果**：返回 `(新增数量, 扫描到的总数)`。普通手动扫描在返回后还会自动调用"重建合集"。

---

## 三、关键特性与注意事项

- **增量、可重复执行**：只新增库里没有的文件；本次扫描过程中也在内存里做去重，防止同一路径被重复添加。反复扫描不会产生重复记录，只是跳过已存在的。
- **不碰磁盘源文件**：扫描只**读取**文件元信息（大小 / 创建时间），**不修改、不移动、不删除**磁盘上的任何文件，也**不解析**书名 / 作者。
- **关键词替换只影响"库内文件名"**：扫描阶段的替换只改写**写入数据库的文件名**，磁盘上的真实文件名保持不变；若想规范化库内书名 / 作者，需在「设置」页配置**解析阶段**规则并重新解析。
- **排除目录规则**：支持"按文件夹名（全层级匹配）"或"相对路径"排除；配置前请确认名称 / 路径正确，否则可能漏扫或误扫。
- **大小写敏感性**：Windows 上用于去重的路径比较**不区分大小写**；Linux / macOS 区分大小写。
- **大目录友好**：文件多时会分批提交（每 50 条一次），进度条走完前请勿重复点击扫描，以免产生大量"跳过"日志与额外开销。

> 扫描只是把文件"登记"进系统；要让文件变得可读、可筛选，还需进行《功能说明》里的**工程解析**；要批量整理冗余，则见**一键清理**与**合集 / 标记重复**。''',
        },
        {
            'doc_key': 'features',
            'title': '功能说明',
            'sort_order': 3,
            'content': '''# 功能说明

## 1. 扫描配置
管理一个或多个待扫描文件夹，每个配置独立保存：名称、目录、文件类型、排除子目录。
支持新增 / 编辑 / 删除；**删除配置会连带删除其下所有扫描结果与解析数据**（仅删数据，不删源文件）。

扫描的具体执行逻辑（校验目录、加载规则、增量去重、排除目录匹配、批量入库等），以及手动扫描与一键清理扫描节点的区别，详见《扫描逻辑详解》。

## 2. 数据列表
表格展示扫描到的文件，支持：
- 关键词模糊搜索（文件名 / 路径 / 小说名 / 作者）
- 按小说名筛选
- 列显示配置（勾选需要展示的字段）
- 行内编辑 `书名 / 作者 / 摘要 / 进度 / 来源`
- 多选（勾选行）用于批量解析与导出

## 3. 工程解析（本地正则）
全部基于文件名与文件内容的**正则规则**，不调用任何大模型或外部接口：
- **文件名解析**：提取 `书名 / 作者 / 进度 / 来源`。
  - 范围：**已选 / 全部（跳过已有）/ 强制重跑全部**。
- **摘要提取**：读取正文开头，截取首章前的简介填入「摘要」。
  - 范围同上。
- 解析在后台异步进行，进度条实时显示成功 / 失败数量。

## 4. 合集模式与重复标记
- **合集模式**：在「数据列表」页点「合集模式」按钮，文件按**小说名**聚合成合集，便于同类归并查看；再次点击退出合集视图。
- **「标记重复」规则**（在**同一合集内、按 `(作者 + 小说名)` 子分组**套用，与一键清理、APP 端完全一致；核心算法见 `backend/dup_logic.py`）：

  | 规则 | 名称 | 判定 |
  | --- | --- | --- |
  | ① | 完全相等去重 | `文件名 + 大小 + 小说名 + 作者 + 进度` 五字段完全一致的**多本**中，保留最新(创建最晚，并列取 id 最大) 一本，**其余全部勾选**。 |
  | ② | 纯数字进度对比 | 同 `(作者+小说名)` 内，若**所有**进度均为纯数字(可含小数、尾随 `%`)，则**进度数字最大者不勾选**，其余纯数字文件全部勾选。 |
  | ③ | 含中文进度 / 完结特例 | 进度含中文(如「完结/连载/断更」) 的**不勾选**；但若同组存在文件名带「完结/完本/全本/全集/完整/全套/全集版」等关键词、且「进度数字最大文件」的大小 **小于** 同组**所有**含中文进度文件的大小，则该「进度数字最大文件」**也要勾选**（存在更完整的完结版，部分进度版冗余应删）。 |
  | ④ | 最大文件不勾选原则 | 已勾选文件中，本 `(作者+小说名)` 组内**唯一**文件大小最大者**不勾选**；若多本并列最大，则按大小无法区分、不据此保护，以免同尺寸重复组被整体保留。 |
  | ⑤ | 完结+N番外 组合排序 | 进度【严格】匹配 `完结+数字番外`（如「完结+3番外」）的文件，在同 `(作者+小说名)` 组内按数字 N 排序：**数字最大者不勾选**，其余勾选（强制勾选，覆盖规则③A 对中文进度的保护）；但被勾选的文件若恰为本组 `(作者+小说名)` 内**文件大小最大者**，则也不勾选。 |

  > 最终勾选集 = (规则①/②/⑤ 勾选集 − 规则①/③/④/⑤ 保护集) ∪ 规则①/③/⑤ 强制勾选集。
  > 规则①的精确重复(=同尺寸)不可能同时是「唯一最大文件」，故强制勾选不会与规则④冲突。

  **举例**：
  - 完全相等去重（规则①）：3 个文件五字段完全一致，创建时间 A(最早)/B/C(最新) → C 不勾选，A、B 勾选（删旧留新）。
  - 纯数字进度对比（规则②）：进度 `12% / 30% / 30%` → 两个 `30%` 并列最大不勾选，`12%` 勾选；若 `12% / 30% / 50%` → `50%` 不勾选，`12%`、`30%` 勾选。
  - 含中文进度 + 完结特例（规则③）：F1 进度 `80%`(2.0MB)、F2 `斗破苍穹完结版.txt` 进度 `完结`(5.0MB)、F3 进度 `完结`(6.0MB) → F2/F3 含中文受保护不勾选；F1 进度最大但比所有含中文文件都小，且组内有「完结」关键词 → F1 强制勾选(删)，F2/F3 保留更完整的完结版。
  - 最大文件保护（规则④）：进度均为 `50% / 50% / 50% / 50%`、大小 1/2/2/4 MB → 进度并列最大全部不勾选，唯一最大 4MB 即便被勾选也保护；本例最终不勾选任何文件。

  - 完结+N番外 组合（规则⑤）：F1 `完结+3番外`(2.0MB)、F2 `完结+5番外`(5.0MB)、F3 `完结`(6.0MB) → M={F1,F2} 按 N 排序，数字最大 F2 不勾选，F1 勾选（除非 F1 恰为本组最大文件）；F3 含中文受规则③保护不勾选。最终 F1 勾选、F2/F3 保留。

- 标记只是**勾选**，不删除；在合集视图下可直接批量清理。

## 5. 导出
- **导出 MD**：选中文件 → 每个文件导出一个 Markdown（含书名 / 作者 / 进度 / 来源 / 摘要），保存到 `exports/`。
- **导出 Excel**：选中或全部 → 导出为 Excel（列：文件名、小说名、作者、进度、来源、摘要）。

## 6. 关键词替换（设置页）
在「设置」页维护替换规则，分两个作用阶段：
- **扫描阶段（scope=scan）**：作用于**文件名**，在入库前改写。
- **解析阶段（scope=parse）**：作用于 **书名 / 作者 / 进度 / 来源**，在解析后规范化。
多条规则按 `sort_order` 从小到大依次生效；可用「启用」开关临时停用，不必删除。

## 7. 删除操作（两种模式）
删除 / 清空均提供两种模式，**务必看清再确认**：
- **仅删除数据（db）**：只删除数据库记录，**保留你的物理源文件**；可后续重新扫描恢复。
- **删除文件（file）**：同时删除数据库记录和**磁盘上的真实源文件**，删除后**不可恢复**。
> 警告：`file` 模式会真正删除你硬盘上的文件，请确认目录无误、重要数据已备份。

普通删除 / 清空都有确认弹窗；勾选「删除文件」后操作较慢（逐文件删除），进度条会显示进度。

## 8. 一键清理（Pipeline）
选定配置后，后台**串行**自动完成全流程：
1. **扫描** → 2. **工程文件名解析（全部）** → 3. **生成合集** → 4. **标记重复** → 5. **清理删除**
- 节点 5 在真正删除前进入「二次确认」状态，前端展示**将被删除的清单**，确认后才执行；取消则放弃删除。
- **单实例约束**：同一时间只能有一个一键清理在跑，新启动会被拒绝（请先取消或等待前一个完成）。
- **取消语义**：取消后「已完成节点保留、不回滚」，后续节点（含删除）跳过。
- 两种删除模式同第 7 节（`db` / `file`），在启动一键清理时选择。
- 流程进度、日志实时显示在「一键清理」页，可随时取消。

## 9. 帮助手册（本模块）
帮助内容由程序内置的默认文档提供，支持 Markdown 渲染。如需自定义，可直接修改数据库 `help_docs` 表对应记录；程序升级时会以内置默认内容同步覆盖。''',
        },
        {
            'doc_key': 'faq',
            'title': '常见问题',
            'sort_order': 4,
            'content': '''# 常见问题

### 解析需要联网吗？
不需要。本系统的解析全部基于**本地工程正则方法**，不调用任何大模型或外部接口，不会产生任何费用。

### 为什么没有「AI 解析」「Token 统计」功能了？
系统已移除全部与大模型 / AI 相关的解析与统计能力，仅保留稳定、免费的本地工程解析，避免密钥配置与费用开销。

### 进度字段为什么是「更78 / 完结613+番外」这种格式？
进度直接从文件名中正则提取，格式取决于原文件名。可在「设置」页添加「解析阶段」关键词替换规则将其规范化，再对相关配置「强制重跑全部」对应解析。

### 标记重复 / 一键清理会删除我的文件吗？
取决于你选择的删除模式：
- **仅删除数据**：不碰源文件，可重新扫描恢复；
- **删除文件**：会真正删除磁盘上的源文件（一键清理删前有「二次确认」清单，普通删除有确认弹窗）。
请务必确认目录正确、重要数据已备份。

### 一键清理中途能取消吗？
可以。取消后已完成的节点（扫描 / 解析 / 合集 / 标记）结果保留，删除不会执行；注意**不回滚**已完成的步骤。

### 数据存在哪里？
数据保存在数据库，具体位置取决于你使用的版本：
- **本地软件（EXE）版本**：使用内置 **SQLite** 数据库，默认位于程序**安装目录**下的 `FileScannerData\\file_scanner.db`（如 `C:\\Users\\<你的用户名>\\AppData\\Local\\Programs\\FileScanner\\FileScannerData\\file_scanner.db`）。日志与导出文件也统一存放于该 `FileScannerData\\` 文件夹内，数据库、日志、导出随安装目录整体迁移。无需安装 MySQL。
- **网页版**：使用 **MySQL** 数据库，默认库名为 `file_scanner_noai`。扫描结果、解析数据、合集、关键词规则、帮助文档等都存在对应表中，可直接查看与维护。

### 解析结果不准怎么办？
- 先确认源文件名是否规范；
- 在「设置」页加「解析阶段」关键词替换规则做规范化；
- 对受影响配置用「强制重跑全部」重新解析以套用新规则；
- 或直接在数据列表里行内编辑修正。''',
        },
    ]
    # 以代码内置内容为准：已存在的 doc_key 更新，缺失的插入（无论表是否为空都同步）。
    for d in docs:
        existing = db.query(HelpDoc).filter(HelpDoc.doc_key == d['doc_key']).first()
        if existing:
            existing.title = d['title']
            existing.content = d['content']
            existing.sort_order = d['sort_order']
        else:
            db.add(HelpDoc(**d))
    db.commit()
    logger.info(f'帮助手册默认文档已同步: {len(docs)} 篇')


# 合集（文件分组）重建进度：config_id -> {'done':已处理扫描记录数, 'total':总记录数, 'phase':'...'}
rebuild_progress = {}


def rebuild_file_groups(config_id: int, db: Session) -> int:
    """重建指定扫描配置的文件分组（合集）预计算结果

    清空该配置下已有的 FileGroup 记录，然后从 scan_results 和 file_metadata
    重新计算分组数据并写入 file_groups 表。

    为支持大库（10w+ 文件）下的进度展示，本函数改为**分批**处理：
    先统计总分组的扫描记录基数，按 scan_results.id 分桶每批 2000 条，
    逐批 DELETE 旧分组 + 重新计算该桶的分组并 UPSERT 写入；每批结束更新
    `rebuild_progress[config_id]` 供前端进度条轮询。

    Args:
        config_id: 扫描配置 ID
        db: 数据库会话（调用方负责提交/关闭；本函数内部每批自行 commit）

    Returns:
        创建/更新的分组数量
    """
    try:
        logger.info(f'[重建分组] 开始: config_id={config_id}')
        t0 = time.perf_counter()
        rebuild_progress[config_id] = {'done': 0, 'total': 0, 'phase': '统计基数'}

        # 0. 统计该配置下的总扫描记录数（进度分母）
        total_rows = db.query(func.count(ScanResult.id)).filter(
            ScanResult.scan_config_id == config_id
        ).scalar() or 0
        rebuild_progress[config_id]['total'] = total_rows
        logger.info(f'[重建分组] config_id={config_id} 共 {total_rows} 条扫描记录')

        # 1. 先删除该配置下所有旧的分组记录（整批一次，数据量小）
        del_t = time.perf_counter()
        del_count = db.query(FileGroup).filter(FileGroup.config_id == config_id).delete(
            synchronize_session=False
        )
        db.commit()
        logger.debug(f'[重建分组] 已删除旧分组 {del_count} 条, 耗时={time.perf_counter() - del_t:.3f}s')

        # 2. 分批（按 scan_results.id 分桶）重算并写入，逐步上报进度
        BATCH = 2000
        if total_rows == 0:
            rebuild_progress[config_id].update({'done': 0, 'phase': '完成'})
            group_count = 0
            logger.info(f'重建文件分组完成(空库): config_id={config_id}, groups=0')
            return 0

        # 取本配置所有 scan_results.id 的最小/最大值，按 id 区间分桶
        id_range = db.query(
            func.min(ScanResult.id), func.max(ScanResult.id)
        ).filter(ScanResult.scan_config_id == config_id).first()
        min_id, max_id = (id_range[0] or 1), (id_range[1] or 0)
        bucket_count = max(1, (max_id - min_id + 1 + BATCH - 1) // BATCH)

        group_count = 0
        processed = 0
        for b in range(bucket_count):
            lo = min_id + b * BATCH
            hi = lo + BATCH - 1
            rebuild_progress[config_id]['phase'] = f'计算分组 {b + 1}/{bucket_count}'
            sql = text("""
                INSERT INTO file_groups (config_id, novel_name, file_count, total_size)
                SELECT
                    sr.scan_config_id,
                    COALESCE(fm.novel_name, '') AS novel_name,
                    COUNT(*) AS file_count,
                    COALESCE(SUM(sr.file_size), 0) AS total_size
                FROM scan_results sr
                LEFT JOIN file_metadata fm ON sr.id = fm.scan_result_id
                WHERE sr.scan_config_id = :config_id
                  AND sr.id BETWEEN :lo AND :hi
                GROUP BY sr.scan_config_id, COALESCE(fm.novel_name, '')
                ON DUPLICATE KEY UPDATE
                    file_count = VALUES(file_count),
                    total_size = VALUES(total_size)
            """) if not IS_SQLITE else text("""
                INSERT INTO file_groups (config_id, novel_name, file_count, total_size)
                SELECT
                    sr.scan_config_id,
                    COALESCE(fm.novel_name, '') AS novel_name,
                    COUNT(*) AS file_count,
                    COALESCE(SUM(sr.file_size), 0) AS total_size
                FROM scan_results sr
                LEFT JOIN file_metadata fm ON sr.id = fm.scan_result_id
                WHERE sr.scan_config_id = :config_id
                  AND sr.id BETWEEN :lo AND :hi
                GROUP BY sr.scan_config_id, COALESCE(fm.novel_name, '')
            """)
            # SQLite 不支持 INSERT...ON DUPLICATE，但本批已先整体 DELETE 旧分组，
            # 且按 id 桶划分后每个桶的分组名集合互不重叠，不会冲突，可直接 INSERT。
            db.execute(sql, {'config_id': config_id, 'lo': lo, 'hi': hi})
            db.commit()
            processed = min(processed + BATCH, total_rows)
            rebuild_progress[config_id]['done'] = processed
            logger.debug(f'[重建分组] 桶 {b + 1}/{bucket_count} 完成, 已处理 {processed}/{total_rows}')

        # 统计创建的分组数
        group_count = db.query(FileGroup).filter(
            FileGroup.config_id == config_id
        ).count()
        rebuild_progress[config_id]['phase'] = '完成'
        logger.info(f'重建文件分组完成: config_id={config_id}, groups={group_count}, '
                    f'总耗时={time.perf_counter() - t0:.1f}s')
        return group_count
    except Exception as e:
        db.rollback()
        rebuild_progress.pop(config_id, None)
        logger.error(f'重建文件分组失败: config_id={config_id}, error={e}', exc_info=True)
        return 0


def get_db():
    """获取数据库会话的依赖注入"""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ===================== FastAPI 应用 =====================
# 获取项目根目录
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 前端静态资源目录：打包为 EXE 后，依据 PyInstaller 的 sys._MEIPASS
# （单文件模式）或可执行文件所在目录（单文件夹模式）定位 frontend 目录。
if getattr(sys, 'frozen', False):
    _BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
else:
    _BASE_DIR = PROJECT_ROOT
FRONTEND_DIR = os.path.join(_BASE_DIR, 'frontend')


@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期：启动时初始化数据库，关闭时清理资源"""
    init_database()
    # 向一键清理流程管理器注入数据库与会话依赖（避免循环 import）
    pipeline_manager.init_pipeline(
        session_factory=SessionLocal,
        pymysql_factory=(None if IS_SQLITE else make_pymysql_conn),
        log_func=add_parse_log,
        rebuild_file_groups_func=rebuild_file_groups,
    )
    logger.info('文件扫描管理系统启动完成')
    yield
    # 关闭时的清理逻辑
    logger.info('文件扫描管理系统正在关闭...')


app = FastAPI(title='文件扫描管理系统', version='1.0.0', lifespan=lifespan)

# CORS配置（收紧为指定来源，禁止完全开放）
CORS_ORIGINS = [o.strip() for o in _os.environ.get(
    'CORS_ORIGINS', 'http://localhost:8000,http://127.0.0.1:8000'
).split(',') if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=False,
    allow_methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'],
    allow_headers=['*'],
)


@app.middleware('http')
async def cache_static_assets(request: Request, call_next):
    """为版本化的静态资源（/static）设置较长缓存，提升前端重复加载性能。
    配合 index.html 中带 ?v= 的资源链接，部署后通过更新版本号即可失效缓存。
    """
    response = await call_next(request)
    if request.url.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=604800, immutable'
    return response


# ===================== 扫描配置 API =====================

@app.get('/api/configs')
def list_configs(db: Session = Depends(get_db)):
    """获取所有扫描配置"""
    configs = db.query(ScanConfig).order_by(ScanConfig.id.desc()).all()
    return [
        {
            'id': c.id,
            'name': c.name or '',
            'folder_path': c.folder_path,
            'file_types': c.file_types,
            'excluded_folders': c.excluded_folders or '',
            'parse_on_scan': bool(c.parse_on_scan),
            'created_at': c.created_at.isoformat() if c.created_at else None,
        }
        for c in configs
    ]


@app.post('/api/configs')
def create_config(data: dict, db: Session = Depends(get_db)):
    """创建扫描配置"""
    parse_on_scan = data.get('parse_on_scan', True)
    config = ScanConfig(
        name=(data.get('name') or '').strip(),
        folder_path=data['folder_path'],
        file_types=data.get('file_types', 'txt'),
        excluded_folders=data.get('excluded_folders', ''),
        parse_on_scan=parse_on_scan if isinstance(parse_on_scan, bool) else str(parse_on_scan).lower() in ('1', 'true', 'yes', 'on'),
    )
    db.add(config)
    db.commit()
    db.refresh(config)
    logger.info(f'创建扫描配置: ID={config.id}, path={config.folder_path}, parse_on_scan={config.parse_on_scan}')
    try:
        from backend.operation_log import log_operation
        log_operation('创建扫描配置', detail=f'{config.name or config.folder_path}',
                      config_id=config.id, path=config.folder_path, parse_on_scan=config.parse_on_scan)
    except Exception:
        pass
    return {'id': config.id, 'message': '创建成功'}


@app.put('/api/configs/{config_id}')
def update_config(config_id: int, data: dict, db: Session = Depends(get_db)):
    """更新扫描配置"""
    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail='配置不存在')

    if 'folder_path' in data:
        config.folder_path = data['folder_path']
    if 'name' in data:
        config.name = (data.get('name') or '').strip()
    if 'file_types' in data:
        config.file_types = data['file_types']
    if 'excluded_folders' in data:
        config.excluded_folders = data.get('excluded_folders', '')
    if 'parse_on_scan' in data:
        v = data['parse_on_scan']
        config.parse_on_scan = v if isinstance(v, bool) else str(v).lower() in ('1', 'true', 'yes', 'on')

    db.commit()
    logger.info(f'更新扫描配置: ID={config_id}, parse_on_scan={config.parse_on_scan}')
    try:
        from backend.operation_log import log_operation
        log_operation('更新扫描配置', detail=f'{config.name or config.folder_path}',
                      config_id=config_id, parse_on_scan=config.parse_on_scan)
    except Exception:
        pass
    return {'message': '更新成功'}


@app.delete('/api/configs/{config_id}')
def delete_config(config_id: int, db: Session = Depends(get_db)):
    """删除扫描配置（同时清理关联的分组、扫描结果和元数据）"""
    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail='配置不存在')

    # 先清理关联表，避免外键约束阻止删除
    try:
        # 1. 删除关联的 file_metadata
        if IS_SQLITE:
            # SQLite 不支持多表 JOIN 的 DELETE 语法，改用子查询
            db.execute(text(
                "DELETE FROM file_metadata WHERE scan_result_id IN "
                "(SELECT id FROM scan_results WHERE scan_config_id = :cid)"
            ), {'cid': config_id})
        else:
            db.execute(text("DELETE fm FROM file_metadata fm "
                            "INNER JOIN scan_results sr ON fm.scan_result_id = sr.id "
                            "WHERE sr.scan_config_id = :cid"), {'cid': config_id})
        # 2. 删除关联的 file_groups
        db.query(FileGroup).filter(FileGroup.config_id == config_id).delete(synchronize_session=False)
        # 4. 删除关联的 scan_results
        db.query(ScanResult).filter(ScanResult.scan_config_id == config_id).delete(synchronize_session=False)
        # 5. 删除扫描配置本身
        db.delete(config)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f'删除扫描配置失败: ID={config_id}, error={e}', exc_info=True)
        raise HTTPException(status_code=500, detail=f'删除失败: {str(e)}')

    logger.info(f'删除扫描配置: ID={config_id}, path={config.folder_path}, file_types={config.file_types}')
    try:
        from backend.operation_log import log_operation
        log_operation('删除扫描配置', detail=f'{config.name or config.folder_path}',
                      config_id=config_id, path=config.folder_path)
    except Exception:
        pass
    return {'message': '删除成功'}


# ===================== 扫描执行 API =====================

@app.post('/api/scan/{config_id}')
def run_scan(config_id: int, db: Session = Depends(get_db)):
    """执行扫描"""
    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail='配置不存在')

    file_types = [ft.strip() for ft in config.file_types.split(',') if ft.strip()]
    excluded = [f.strip() for f in (config.excluded_folders or '').split(',') if f.strip()]

    logger.info(f'开始扫描: config_id={config_id}, path={config.folder_path}, file_types={config.file_types}')

    # 初始化进度（包含起始时间和计数）
    scanning_progress[config_id] = {'count': 0, 'start_time': datetime.now()}
    try:
        new_count, total = scan_files(
            db=db,
            config_id=config.id,
            folder_path=config.folder_path,
            file_types=file_types,
            excluded_folders=excluded,
            progress_dict=scanning_progress,
        )
        # 标记完成
        elapsed = round((datetime.now() - scanning_progress[config_id]['start_time']).total_seconds())
        scanning_progress[config_id]['done'] = True
        scanning_progress[config_id]['elapsed'] = elapsed
        logger.info(f'扫描完成: config_id={config_id}, 新增={new_count}, 总计={total}, 耗时={elapsed}s')
        try:
            from backend.operation_log import log_operation
            log_operation('执行扫描', detail=f'{config.name or config.folder_path}：共{total}个（新增{new_count}个），耗时{elapsed}s',
                          config_id=config_id, total=total, new=new_count, elapsed=elapsed,
                          parse_on_scan=bool(config.parse_on_scan))
        except Exception:
            pass
        # 注意：重建文件分组（合集计算）不再在此同步执行，以避免大库下阻塞
        # 扫描请求直至重建结束。改为扫描返回后由前端显式调用
        # POST /api/groups/rebuild/{config_id} 后台异步重建，并轮询进度条。
        rebuild_progress.pop(config_id, None)

        # 扫描时同步进行工程类解析（默认开启）：扫描完成后自动启动「工程文件名解析」
        # 后台任务，前端拿到 parse_task_id 后轮询解析进度，解析完成再重建合集，
        # 使合集/标记重复立即可用，无需用户再手动点「文件名解析」。
        parse_task_id = None
        if getattr(config, 'parse_on_scan', True) and total > 0:
            try:
                parse_task_id = _start_parse_task(
                    parse_file_names_regex_only, f'扫描同步工程文件名解析 ({total}个)',
                    SessionLocal, concurrency=8, force=False,
                    config_id=config_id, forward_config_id=True,
                    pymysql_factory=(None if IS_SQLITE else make_pymysql_conn),
                )
                logger.info(f'扫描后自动启动工程文件名解析: config_id={config_id}, task_id={parse_task_id}, file_count={total}')
            except Exception:
                logger.exception('扫描后自动启动工程文件名解析失败')

        msg = f'扫描完成，共{total}个文件（新增{new_count}个，耗时{elapsed}秒）'
        if parse_task_id:
            msg += '，正在同步进行工程类解析...'
        return {'new_count': new_count, 'total': total, 'parse_task_id': parse_task_id, 'message': msg}
    except FileNotFoundError as e:
        scanning_progress.pop(config_id, None)
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        scanning_progress.pop(config_id, None)
        logger.exception('扫描执行出错')
        raise HTTPException(status_code=500, detail=f'扫描出错: {str(e)}')


# ===================== 扫描进度 API =====================

@app.get('/api/scan-progress/{config_id}')
def get_scan_progress(config_id: int):
    state = scanning_progress.get(config_id)
    if state is None:
        return {'config_id': config_id, 'count': -1, 'done': True, 'elapsed': 0}
    count = state.get('count', 0)
    done = state.get('done', False)
    start_time = state.get('start_time')
    if start_time and not done:
        elapsed = round((datetime.now() - start_time).total_seconds())
    else:
        elapsed = state.get('elapsed', 0)
    return {'config_id': config_id, 'count': count, 'done': done, 'elapsed': elapsed}


@app.get('/api/rebuild-progress/{config_id}')
def get_rebuild_progress(config_id: int):
    """获取合集（文件分组）重建进度，供前端进度条轮询。

    扫描结束的“重建合集”与一键清理的「生成合集」节点都会写
    `rebuild_progress[config_id]`。若当前无重建任务在进行，返回 done=True。
    """
    state = rebuild_progress.get(config_id)
    if state is None:
        return {'config_id': config_id, 'done': True, 'done_count': 0,
                'total': 0, 'phase': '', 'percent': 100}
    done_count = state.get('done', 0)
    total = state.get('total', 0)
    percent = round((done_count * 100 / total), 1) if total > 0 else 100
    return {
        'config_id': config_id,
        'done': state.get('phase') == '完成',
        'done_count': done_count,
        'total': total,
        'phase': state.get('phase', ''),
        'percent': percent,
    }


@app.post('/api/groups/rebuild/{config_id}')
def api_rebuild_groups(config_id: int, db: Session = Depends(get_db)):
    """后台异步重建某配置的合集（文件分组），并实时上报进度。

    扫描完成 / 清空 / 删除后由前端显式调用；前端通过
    GET /api/rebuild-progress/{config_id} 轮询进度条。
    同一配置已有重建任务进行中时，返回已在运行，避免重复触发。
    """
    if rebuild_progress.get(config_id) is not None and rebuild_progress[config_id].get('phase') != '完成':
        return {'message': '该配置正在重建合集，请稍候', 'running': True}

    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail='扫描配置不存在')

    def _bg():
        bg_db = SessionLocal()
        try:
            rebuild_file_groups(config_id, bg_db)
        finally:
            bg_db.close()
        # 重建完成稍作保留，便于前端最后一次轮询拿到 100%
        import threading as _th
        def _clear():
            rebuild_progress.pop(config_id, None)
        _th.Timer(2.0, _clear).start()

    threading.Thread(target=_bg, daemon=True).start()
    logger.info(f'已触发后台重建合集: config_id={config_id}')
    return {'message': '已触发后台重建合集', 'running': False}


# ===================== 扫描结果 API =====================

# ===================== 小说名筛选 API =====================

@app.get('/api/novel-names')
def list_novel_names(
    config_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=5000),
    db: Session = Depends(get_db),
):
    """获取小说名列表及其重复数量（含空白项），支持分页"""
    # 统计空白小说名的数量
    has_empty_filter = not search or '__empty__' in search.lower()
    empty_count = 0
    if page == 1 and has_empty_filter:
        empty_query = db.query(func.count(ScanResult.id)).outerjoin(
            FileMetadata, ScanResult.id == FileMetadata.scan_result_id
        )
        if config_id:
            empty_query = empty_query.filter(ScanResult.scan_config_id == config_id)
        empty_query = empty_query.filter(
            (FileMetadata.novel_name == '') | (FileMetadata.novel_name.is_(None))
        )
        empty_count = empty_query.scalar() or 0

    # 获取非空小说名列表
    query = db.query(
        FileMetadata.novel_name,
        func.count(ScanResult.id).label('count'),
    ).join(
        ScanResult, ScanResult.id == FileMetadata.scan_result_id
    )

    if config_id:
        query = query.filter(ScanResult.scan_config_id == config_id)

    if search:
        query = query.filter(FileMetadata.novel_name.like(f'%{search}%'))

    query = query.filter(FileMetadata.novel_name != '')
    query = query.group_by(FileMetadata.novel_name)
    query = query.order_by(func.count(ScanResult.id).desc())

    # 计算总数
    total = query.count()

    # 分页
    results = query.offset((page - 1) * page_size).limit(page_size).all()
    items = [
        {'novel_name': r.novel_name, 'count': r.count}
        for r in results
    ]

    # 空白项放在第一页最前面
    if page == 1 and empty_count > 0:
        items.insert(0, {'novel_name': '__empty__', 'count': empty_count})
        total += 1  # 空白项计入总数
        items = items[:page_size]  # 确保不超出一页

    total_pages = max(1, (total + page_size - 1) // page_size)
    return {
        'items': items,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
    }


# ===================== 扫描结果 API =====================

@app.delete('/api/results')
def delete_results(ids: List[int] = Query(...), db: Session = Depends(get_db)):
    """批量删除扫描结果（仅删数据库记录，不删源文件）"""
    if not ids:
        raise HTTPException(status_code=400, detail='请选择要删除的记录')

    records = db.query(ScanResult).filter(ScanResult.id.in_(ids)).all()
    if not records:
        return {'deleted': 0, 'message': '未找到对应记录'}

    config_id = records[0].scan_config_id
    # 先删关联的 file_metadata，再删 scan_results（ORM 方式，避免 IN 参数绑定差异）
    db.query(FileMetadata).filter(FileMetadata.scan_result_id.in_(ids)).delete(synchronize_session=False)
    deleted_count = db.query(ScanResult).filter(ScanResult.id.in_(ids)).delete(synchronize_session=False)
    db.commit()

    # 重建文件分组
    rebuild_file_groups(config_id, db)
    logger.info(f'仅删除数据库记录: 成功={deleted_count}, ids_count={len(ids)}')
    try:
        from backend.operation_log import log_operation
        log_operation('删除记录（仅数据库）', detail=f'成功删除{deleted_count}条记录',
                      config_id=config_id, deleted=deleted_count)
    except Exception:
        pass
    return {'deleted': deleted_count, 'message': f'成功删除{deleted_count}条记录'}


@app.delete('/api/results/with-files')
def delete_results_with_files(ids: List[int] = Query(...), db: Session = Depends(get_db)):
    """批量删除扫描结果并删除源文件（先删源文件，验证成功后，再删数据库记录）"""
    if not ids:
        raise HTTPException(status_code=400, detail='请选择要删除的记录')

    records = db.query(ScanResult).filter(ScanResult.id.in_(ids)).all()
    if not records:
        return {'deleted': 0, 'failed_files': [], 'message': '未找到对应记录'}

    # 获取该配置对应的扫描目录，用于路径安全校验
    config_id = records[0].scan_config_id
    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    scan_root = os.path.abspath(config.folder_path) if config else ''

    deleted_count = 0
    failed_files = []
    deleted_ids = []  # 收集成功删除的ID，批量提交

    for record in records:
        try:
            file_path = record.file_path

            # 路径安全检查：只删除扫描目录下的文件
            if scan_root and not _is_path_under_root(scan_root, file_path):
                logger.warning(f'跳过非扫描目录文件: {file_path}')
                failed_files.append(file_path)
                continue

            # 先删除物理文件（Windows 下可能因杀软/缓存导致删除后仍短暂存在，重试最多3次）
            if os.path.exists(file_path):
                for attempt in range(3):
                    try:
                        os.remove(file_path)
                    except Exception:
                        pass
                    time.sleep(0.1)
                    if not os.path.exists(file_path):
                        break

            # 检查文件是否真的已删除
            if os.path.exists(file_path):
                logger.error(f'文件删除后仍然存在: {file_path}')
                failed_files.append(file_path)
                continue

            # 文件已物理删除，记录ID待批量清理数据库
            deleted_ids.append(record.id)
            deleted_count += 1
        except Exception as e:
            logger.error(f'删除失败: id={record.id}, path={record.file_path}, {e}')
            failed_files.append(record.file_path)
            # 文件未删除成功，不回滚DB（因为还没删DB记录）

    # 批量删除数据库记录（从循环中移出，一次性提交以大幅减少事务次数）
    if deleted_ids:
        db.query(ScanResult).filter(ScanResult.id.in_(deleted_ids)).delete(
            synchronize_session=False
        )
        db.commit()

    id_preview = ','.join(str(i) for i in ids[:5]) + (f'...({len(ids)}个)' if len(ids) > 5 else '')
    logger.info(f'批量删除文件及记录: 成功={deleted_count}, 失败={len(failed_files)}, ids=[{id_preview}]')
    # 重建文件分组
    rebuild_file_groups(config_id, db)
    fail_suffix = f'，{len(failed_files)}个失败' if failed_files else ''
    try:
        from backend.operation_log import log_operation
        log_operation(
            action='删除记录（含文件）',
            detail=f'成功删除{deleted_count}个文件{fail_suffix}',
            config_id=config_id,
            requested=len(ids),
            deleted=deleted_count,
            failed=len(failed_files),
        )
    except Exception as _ole:
        logger.warning(f'操作日志写入失败: {_ole}')

    return {
        'deleted': deleted_count,
        'failed_files': failed_files,
        'message': f'成功删除{deleted_count}个文件{fail_suffix}',
    }


@app.delete('/api/configs/{config_id}/clear')
def clear_config_results(config_id: int, with_files: bool = Query(False), db: Session = Depends(get_db)):
    """清空指定扫描配置下的所有扫描结果（如需删源文件，先删文件，验证成功后，再删数据库记录）"""
    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail='扫描配置不存在')

    records = db.query(ScanResult).filter(ScanResult.scan_config_id == config_id).all()
    if not records:
        return {'deleted': 0, 'message': '没有需要清空的数据'}

    scan_root = os.path.abspath(config.folder_path)

    total_count = len(records)
    deleted_files = 0
    failed_files = []
    successful_ids = []

    if with_files:
        # 先逐条删除物理文件，收集成功删除的ID
        for record in records:
            try:
                file_path = record.file_path
                # 路径安全检查
                if not _is_path_under_root(scan_root, file_path):
                    logger.warning(f'跳过非扫描目录文件: {file_path}')
                    failed_files.append(file_path)
                    continue

                # 先删除物理文件
                if os.path.exists(file_path):
                    os.remove(file_path)

                # 检查文件是否真的已删除
                if os.path.exists(file_path):
                    logger.error(f'文件删除后仍然存在: {file_path}')
                    failed_files.append(file_path)
                    continue

                # 文件删除成功
                deleted_files += 1
                successful_ids.append(record.id)
            except Exception as e:
                logger.warning(f'删除源文件失败: {record.file_path}, {e}')
                failed_files.append(record.file_path)

        # 再批量删除数据库记录（仅删除已成功删除源文件的记录）
        if successful_ids:
            db.query(ScanResult).filter(ScanResult.id.in_(successful_ids)).delete(
                synchronize_session=False
            )
            db.commit()

        count = len(successful_ids)
    else:
        # 不删源文件，直接批量删除数据库记录
        count = total_count
        db.query(ScanResult).filter(ScanResult.scan_config_id == config_id).delete(
            synchronize_session=False
        )
        db.commit()

    # 重建文件分组
    rebuild_file_groups(config_id, db)

    log_msg = f'清空扫描配置[{config_id}]结果: {count}条'
    if with_files:
        log_msg += f', 删除源文件{deleted_files}个'
    if failed_files:
        log_msg += f', {len(failed_files)}个失败'
    logger.info(f'{log_msg}, path={config.folder_path}')

    fail_suffix = f'，{len(failed_files)}个删除失败' if failed_files else ''
    try:
        from backend.operation_log import log_operation
        log_operation(
            action='清空扫描配置',
            detail=f'成功清空{count}条记录{fail_suffix}',
            config_id=config_id,
            with_files=with_files,
            deleted=count,
            deleted_files=deleted_files,
            failed=len(failed_files),
        )
    except Exception as _ole:
        logger.warning(f'操作日志写入失败: {_ole}')

    return {
        'deleted': count,
        'deleted_files': deleted_files,
        'message': f'成功清空{count}条记录{fail_suffix}',
    }


@app.get('/api/results')
def list_results(
    config_id: Optional[int] = Query(None),
    sort_by: str = Query('id'),
    sort_order: str = Query('desc'),
    search: Optional[str] = Query(None),
    novel_names: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    group_by_file_name: bool = Query(False),
    min_group_count: int = Query(0, ge=0),
    max_group_count: Optional[int] = Query(None, ge=0),
    exclude_names: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """获取扫描结果列表（分页），支持合集模式"""
    base_query = db.query(
        ScanResult.id,
        ScanResult.file_name,
        ScanResult.file_size,
        ScanResult.file_path,
        ScanResult.created_date,
        FileMetadata.novel_name,
        FileMetadata.author,
        FileMetadata.summary,
        FileMetadata.progress,
        FileMetadata.source,
        ScanResult.scan_config_id,
    ).outerjoin(
        FileMetadata, ScanResult.id == FileMetadata.scan_result_id
    )

    if config_id:
        base_query = base_query.filter(ScanResult.scan_config_id == config_id)

    if search:
        like_pat = f'%{search}%'
        from sqlalchemy import or_
        base_query = base_query.filter(or_(
            ScanResult.file_name.like(like_pat),
            ScanResult.file_path.like(like_pat),
            FileMetadata.novel_name.like(like_pat),
            FileMetadata.author.like(like_pat),
        ))

    if novel_names:
        names_list = [n.strip() for n in novel_names.split(',') if n.strip()]
        if names_list:
            has_empty = '__empty__' in names_list
            non_empty_names = [n for n in names_list if n != '__empty__']
            from sqlalchemy import or_
            conditions = []
            if has_empty:
                conditions.append(
                    (FileMetadata.novel_name == '') | (FileMetadata.novel_name.is_(None))
                )
            if non_empty_names:
                conditions.append(FileMetadata.novel_name.in_(non_empty_names))
            if conditions:
                base_query = base_query.filter(or_(*conditions))

    if exclude_names:
         exclude_list = [n.strip() for n in exclude_names.split(',') if n.strip()]
         if exclude_list:
             base_query = base_query.filter(
                 func.coalesce(FileMetadata.novel_name, '').notin_(exclude_list)
             )

    if group_by_file_name:
        # 合集模式：按小说名分组（未解析的放在最后，其余按出现次数降序、作者重复数降序、总大小降序）
        from sqlalchemy import func as sa_func
        from collections import OrderedDict
        import math

        # 第 1 次查询：获取分组列表及每组的统计
        count_sub = base_query.subquery()
        group_expr = sa_func.coalesce(count_sub.c.novel_name, '')
        count_expr = sa_func.count()
        total_size_expr = sa_func.sum(count_sub.c.file_size)
        author_dup_expr = sa_func.count(count_sub.c.author) - sa_func.count(sa_func.distinct(count_sub.c.author))
        group_query = db.query(
            group_expr.label('group_name'),
            count_expr.label('name_count'),
            total_size_expr.label('total_size'),
        ).group_by(
            group_expr
        )
        if min_group_count > 0:
            group_query = group_query.having(count_expr >= min_group_count)
        if max_group_count is not None:
            group_query = group_query.having(count_expr <= max_group_count)
        group_query = group_query.order_by(
            count_expr.desc(),
            author_dup_expr.desc(),
            total_size_expr.desc(),
        )

        total_groups = group_query.count()
        total_pages = max(1, math.ceil(total_groups / page_size))
        page_groups_raw = group_query.offset((page - 1) * page_size).limit(page_size).all()
        # Python 层排序：有小说名的在前（按 count desc 已排好），未解析的放最后
        empty = [g for g in page_groups_raw if not g.group_name]
        non_empty = [g for g in page_groups_raw if g.group_name]
        page_groups = non_empty + empty
        page_group_names = [g.group_name for g in page_groups]

        groups = []
        all_item_ids = []

        if page_group_names:
            # 第 2 次查询：一次性获取当前页所有分组下的全部记录
            # 加总量上限保护，避免单页分组数 × 每组文件数过大（如 100×500）时一次取数万行致内存/网络压力
            items_query = base_query.filter(
                sa_func.coalesce(FileMetadata.novel_name, '').in_(page_group_names)
            ).order_by(
                sa_func.coalesce(FileMetadata.novel_name, ''),
                ScanResult.file_size.desc(),
            ).limit(20000)

            name_lookup = {g.group_name: g for g in page_groups}
            raw_by_name = OrderedDict()
            for r in items_query.all():
                nname = r.novel_name or ''
                if nname not in raw_by_name:
                    raw_by_name[nname] = []
                raw_by_name[nname].append(r)

            for gname in page_group_names:
                g = name_lookup[gname]
                display_name = gname if gname else '未解析'
                raw_items = raw_by_name.get(gname, [])
                built_items = [_build_result_item(r) for r in raw_items]
                groups.append({
                    'novel_name': display_name,
                    'count': g.name_count,
                    'total_size': g.total_size,
                    'items': built_items,
                })
                for item in built_items:
                    all_item_ids.append(item['id'])

        valid_ids = set(all_item_ids)
        return {
            'groups': groups,
            'total_groups': total_groups,
            'total': sum(g['count'] for g in groups),
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'valid_ids': list(valid_ids),
        }
    else:
        # 普通模式：原有排序逻辑
        sort_map = {
            'id': ScanResult.id,
            'file_name': ScanResult.file_name,
            'file_size': ScanResult.file_size,
            'file_path': ScanResult.file_path,
            'created_date': ScanResult.created_date,
            'novel_name': FileMetadata.novel_name,
            'author': FileMetadata.author,
            'summary': FileMetadata.summary,
            'progress': FileMetadata.progress,
            'source': FileMetadata.source,
        }
        sort_order = sort_order if sort_order in ('asc', 'desc') else 'desc'
        order_col = sort_map.get(sort_by, ScanResult.id)
        if sort_order == 'desc':
            order_col = order_col.desc()
        query = base_query.order_by(order_col)

        # 分页
        total = query.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        results = query.offset((page - 1) * page_size).limit(page_size).all()

        items = [
            _build_result_item(r)
            for r in results
        ]

        return {
            'items': items,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
        }


# ===================== 文件分组（合集）API =====================

@app.get('/api/groups')
def list_groups(
    config_id: int = Query(...),
    min_count: int = Query(0, ge=0),
    max_count: Optional[int] = Query(None, ge=0),
    exclude_names: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=2000),
    db: Session = Depends(get_db),
):
    """获取文件分组（合集）列表，基于预计算的 file_groups 表"""
    from collections import OrderedDict

    # 1. 构建分组查询
    group_query = db.query(FileGroup).filter(FileGroup.config_id == config_id)

    if min_count > 0:
        group_query = group_query.filter(FileGroup.file_count >= min_count)
    if max_count is not None:
        group_query = group_query.filter(FileGroup.file_count <= max_count)

    if exclude_names:
        exclude_list = [n.strip() for n in exclude_names.split(',') if n.strip()]
        if exclude_list:
            group_query = group_query.filter(FileGroup.novel_name.notin_(exclude_list))

    # 子查询：计算每个分组的作者重复数量（COUNT(author) - COUNT(DISTINCT author)）
    author_dup_sub = db.query(
        func.coalesce(FileMetadata.novel_name, '').label('group_name'),
        (func.count(FileMetadata.author) - func.count(func.distinct(FileMetadata.author))).label('author_dup_cnt')
    ).outerjoin(
        ScanResult, FileMetadata.scan_result_id == ScanResult.id
    ).filter(
        ScanResult.scan_config_id == config_id
    ).group_by(
        func.coalesce(FileMetadata.novel_name, '')
    ).subquery()

    group_query = group_query.outerjoin(
        author_dup_sub, FileGroup.novel_name == author_dup_sub.c.group_name
    )

    # 总数（分页前）
    total_groups = group_query.count()

    # 排序：file_count DESC, author_dup DESC, total_size DESC，但空 novel_name 放最后
    group_query = group_query.order_by(
        (FileGroup.novel_name == text("''")).asc(),
        FileGroup.file_count.desc(),
        func.coalesce(author_dup_sub.c.author_dup_cnt, 0).desc(),
        FileGroup.total_size.desc(),
    )

    # 分页
    page_groups = group_query.offset((page - 1) * page_size).limit(page_size).all()

    # 2. 批量查询当前页所有分组的 items
    group_names = [g.novel_name for g in page_groups]
    items_by_name = OrderedDict()
    if group_names:
        base_query = db.query(
            ScanResult.id,
            ScanResult.file_name,
            ScanResult.file_size,
            ScanResult.file_path,
            ScanResult.created_date,
            FileMetadata.novel_name,
            FileMetadata.author,
            FileMetadata.summary,
            FileMetadata.progress,
            FileMetadata.source,
            ScanResult.scan_config_id,
        ).outerjoin(
            FileMetadata, ScanResult.id == FileMetadata.scan_result_id
        ).filter(
            ScanResult.scan_config_id == config_id,
            func.coalesce(FileMetadata.novel_name, '').in_(group_names),
        ).order_by(
            ScanResult.file_size.desc(),
        ).limit(20000)

        for r in base_query.all():
            nname = r.novel_name or ''
            if nname not in items_by_name:
                items_by_name[nname] = []
            items_by_name[nname].append(r)

    # 3. 组装响应
    groups_data = []
    total_items = 0
    for g in page_groups:
        raw_items = items_by_name.get(g.novel_name, [])
        built_items = [_build_result_item(r) for r in raw_items]
        total_items += len(built_items)
        groups_data.append({
            'novel_name': g.novel_name if g.novel_name else '未解析',
            'file_count': g.file_count,
            'total_size': g.total_size,
            'items': built_items,
        })

    return {
        'groups': groups_data,
        'total': total_items,
        'total_groups': total_groups,
        'page': page,
        'page_size': page_size,
    }


@app.post('/api/groups/select-duplicates')
def select_duplicates(
    config_id: int = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=2000),
    min_count: int = Query(0, ge=0),
    max_count: Optional[int] = Query(None, ge=0),
    exclude_names: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """
    合集模式下“标记重复”（待删勾选）的完整规则，所有判定在【同一合集】内、
    按 (作者 + 小说名) 子分组进行（算法见 backend/dup_logic.py）：

    规则 1（完全相等去重）：(文件名 + 大小 + 小说名 + 作者 + 进度) 五字段完全一致，
        且同组 >= 2 本时，最新(创建时间最晚，并列取 id 最大)的不勾选，其余全部勾选。
    规则 2（纯数字进度对比）：同 (作者+小说名) 内，若所有进度均为纯数字，则
        进度数字最大的不勾选，其余纯数字文件全部勾选。
    规则 3（含中文进度 / 完结特例）：
        - 进度含中文的，不勾选（保护“完结/连载”等状态文件）；
        - 若同组存在文件名带『完结』等关键词、且“进度数字最大文件”的大小
          小于同组所有含中文进度文件的大小时，该“进度数字最大文件”也要勾选
          （说明有更完整的完结版，部分进度版冗余）。
    规则 4（最大文件不勾选原则）：已勾选的文件若为本 (作者+小说名) 组内文件大小最大者，
        则不勾选。

    返回所有合集（按筛选条件）累计应勾选的文件 id 列表。
    """

    # 1. 先按筛选条件，一次性取出“需要参与判定”的合集名集合（一次查询，
    #    避免逐页重复执行昂贵的合集排序子查询）
    name_query = db.query(FileGroup.novel_name).filter(FileGroup.config_id == config_id)
    if min_count > 0:
        name_query = name_query.filter(FileGroup.file_count >= min_count)
    if max_count is not None:
        name_query = name_query.filter(FileGroup.file_count <= max_count)
    if exclude_names:
        exclude_list = [n.strip() for n in exclude_names.split(',') if n.strip()]
        if exclude_list:
            name_query = name_query.filter(FileGroup.novel_name.notin_(exclude_list))
    valid_names = set(n[0] for n in name_query.all())
    collections_processed = len(valid_names)

    if not valid_names:
        return {
            'ids_to_check': [],
            'summary': {
                'groups_processed': 0,
                'subgroups_with_duplicates': 0,
                'total_checked': 0,
            }
        }

    # 2. 仅一次查询取出该 config 下全部条目，再按 valid_names 在内存中过滤
    #    （避免逐页重复执行昂贵的合集排序子查询）
    items = db.query(
        ScanResult.id,
        ScanResult.file_name,
        ScanResult.file_size,
        func.coalesce(FileMetadata.novel_name, '').label('novel_name'),
        FileMetadata.author,
        FileMetadata.progress,
        ScanResult.created_date,
    ).outerjoin(
        FileMetadata, ScanResult.id == FileMetadata.scan_result_id
    ).filter(
        ScanResult.scan_config_id == config_id,
    ).all()

    rows = [{
        'id': it.id,
        'file_name': it.file_name or '',
        'file_size': it.file_size or 0,
        'novel_name': it.novel_name or '',
        'author': it.author or '',
        'progress': it.progress or '',
        'created_date': it.created_date,
    } for it in items if (it.novel_name or '') in valid_names]

    from backend.dup_logic import compute_duplicate_ids
    all_ids, subgroups_with_dups, dup_detail_lines = compute_duplicate_ids(rows)

    # 3. 操作日志：本次“标记重复”的判定结果（调试用，可一键复制）
    try:
        from backend.operation_log import log_operation, log_block
        log_operation(
            '标记重复', level='INFO',
            config_id=config_id, page=page,
            scanned_collections=collections_processed,
            duplicate_collections=subgroups_with_dups,
            total_checked=len(all_ids),
        )
        if dup_detail_lines:
            log_block('标记重复-重复组明细', dup_detail_lines)
    except Exception as log_ex:
        logger.warning(f"写操作日志失败（不影响主流程）: {log_ex}")

    return {
        'ids_to_check': all_ids,
        'summary': {
            'groups_processed': collections_processed,
            'subgroups_with_duplicates': subgroups_with_dups,
            'total_checked': len(all_ids),
        }
    }


# ===================== 操作日志 / 调试日志（一键复制） =====================

@app.get('/api/operation-logs')
def api_operation_logs(limit: int = 2000):
    """读取 operation.log 中的结构化操作日志（标记重复、清理等），供前端展示与一键复制。"""
    try:
        from backend.operation_log import get_operation_logs
        return {'logs': get_operation_logs(limit=limit)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'读取操作日志失败: {e}')


@app.get('/api/pipeline-logs')
def api_pipeline_logs():
    """读取最近一次一键清理的流水线运行日志（ParseLog 结构化记录），供前端调试与一键复制；无运行记录时返回空数组。"""
    try:
        from backend import pipeline
        logs = pipeline.get_logs()
        return {'logs': logs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'读取调试日志失败: {e}')


@app.get('/api/app-logs')
def api_app_logs(lines: int = 2000):
    """读取运行日志 app.log（调试/排查用，含 DEBUG/INFO 等运行期信息）的最新内容，供独立日志页面展示与一键复制。"""
    try:
        from backend.logger import LOG_FILE
        if not os.path.exists(LOG_FILE):
            return {'logs': []}
        with open(LOG_FILE, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().splitlines()
        tail = content[-lines:] if lines and lines > 0 else content
        return {'logs': tail}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'读取运行日志失败: {e}')


class ResultUpdate(BaseModel):
    """行内编辑更新模型"""
    novel_name: Optional[str] = None
    author: Optional[str] = None
    summary: Optional[str] = None
    progress: Optional[str] = None
    source: Optional[str] = None


@app.put('/api/results/{result_id}')
def update_result(result_id: int, data: ResultUpdate, db: Session = Depends(get_db)):
    """更新单行扫描结果（FileMetadata字段）"""
    try:
        result = db.query(ScanResult).filter(ScanResult.id == result_id).first()
        if not result:
            raise HTTPException(status_code=404, detail='记录不存在')

        meta = db.query(FileMetadata).filter(FileMetadata.scan_result_id == result_id).first()
        if not meta:
            meta = FileMetadata(scan_result_id=result_id)
            db.add(meta)

        update_fields = data.model_dump(exclude_none=True)
        for key, val in update_fields.items():
            setattr(meta, key, val)

        db.commit()
        logger.info(f'编辑保存: result_id={result_id}, fields={list(update_fields.keys())}')
        return {'message': '保存成功', 'id': result_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.exception(f'编辑保存失败: result_id={result_id}')
        raise HTTPException(status_code=500, detail=f'保存失败: {str(e)}')


# ===================== AI连接测试 API =====================

def _start_parse_task(parse_func, description, db_session_factory, *args, config_id: Optional[int] = None, forward_config_id: bool = False, **kwargs):
    """
    在后台监督线程中运行解析任务，返回 task_id。

    说明（线程 vs 进程）：
        该监督函数本身仅做「编排/进度聚合/数据库写入协调」，不做 CPU 密集计算，
        且需与 FastAPI 请求线程共享内存中的任务进度表（parse_tasks / parse_logs），
        因此保留为后台线程。真正的 CPU 密集解析已全部下沉到各解析函数内部的
        ProcessPoolExecutor（多进程）执行，从而绕过 GIL 实现并行。

    Args:
        parse_func: 解析函数 (parse_file_names_regex_only / parse_file_summary_regex_only)
        description: 任务描述
        db_session_factory: SessionLocal
        config_id: 可选，扫描配置ID，用于解析完成后重建文件分组
        forward_config_id: 为 True 时把 config_id 透传给 parse_func（用于「解析全部」
                    走 config_id 分页流式处理，避免超大 IN 子句阻塞）
        *args, **kwargs: 传递给 parse_func（不包含 db 和 progress_callback）
    """
    if forward_config_id and config_id is not None:
        kwargs['config_id'] = config_id
    task_id = str(uuid.uuid4())[:8]

    cancel_event = threading.Event()
    with parse_cancel_lock:
        parse_cancel_events[task_id] = cancel_event

    with parse_tasks_lock:
        parse_tasks[task_id] = {
            'status': 'running',
            'total': 0,
            'processed': 0,
            'success': 0,
            'failed': 0,
            'description': description,
            'result': None,
            'started_at': datetime.now(),
        }

    def cancel_check():
        return cancel_event.is_set()

    def _background_run():
        db = db_session_factory()
        task_start = time.perf_counter()
        logger.info(f'解析后台任务开始执行 [{description}]: task_id={task_id}')
        try:
            def progress_callback(processed, total, success, failed):
                with parse_tasks_lock:
                    parse_tasks[task_id].update({
                        'total': total,
                        'processed': processed,
                        'success': success,
                        'failed': failed,
                    })

            result = parse_func(db, *args, **kwargs, progress_callback=progress_callback,
                                cancel_check=cancel_check, db_session_factory=db_session_factory,
                                log_callback=lambda msg: add_parse_log(task_id, msg))

            # 解析完成后，将内存中的日志批量写入数据库（热循环期间不写DB，避免阻塞主线程）
            try:
                with parse_logs_lock:
                    logs_to_flush = parse_logs.get(task_id, [])[-500:]
                if logs_to_flush:
                    db_flush = db_session_factory()
                    try:
                        for entry in logs_to_flush:
                            db_flush.add(ParseLog(task_id=task_id, message=entry['msg'], level=entry['level']))
                        db_flush.commit()
                    finally:
                        db_flush.close()
                    logger.debug(f'已批量写入 {len(logs_to_flush)} 条解析日志到数据库: task_id={task_id}')
            except Exception as e:
                logger.warning(f'批量写入解析日志到数据库失败: {e}')
            # 检查是否被取消
            if cancel_check():
                with parse_tasks_lock:
                    parse_tasks[task_id].update({
                        'status': 'cancelled',
                        'total': result.get('total', 0) if result else 0,
                        'processed': result.get('total', 0) if result else 0,
                        'success': result.get('success', 0) if result else 0,
                        'failed': result.get('failed', 0) if result else 0,
                        'result': result,
                        'message': '任务已被用户取消',
                        'ended_at': datetime.now(),
                    })
                logger.info(f'解析后台任务被取消 [{description}]: task_id={task_id}, 耗时={time.perf_counter() - task_start:.1f}s')
            else:
                with parse_tasks_lock:
                    parse_tasks[task_id].update({
                        'status': 'done',
                        'total': result.get('total', 0),
                        'processed': result.get('total', 0),
                        'success': result.get('success', 0),
                        'failed': result.get('failed', 0),
                        'result': result,
                        'ended_at': datetime.now(),
                    })
                logger.info(f'解析函数执行完毕 [{description}]: task_id={task_id}, total={result.get("total", 0)}, '
                            f'success={result.get("success", 0)}, failed={result.get("failed", 0)}, '
                            f'耗时={time.perf_counter() - task_start:.1f}s')

            # ------------------------------------------------------------
            # 优化: 解析函数结束后，关闭旧会话以释放数据库连接
            # （解析过程中 pymysql 已独立完成写入，旧会话仅用于最初的查询）
            # 重建文件分组使用新会话，避免长时闲置的连接超时问题。
            # ------------------------------------------------------------
            t_cleanup = time.perf_counter()
            db.close()
            logger.info(f'后台任务关闭旧数据库会话, 耗时={time.perf_counter() - t_cleanup:.3f}s')

            # 解析完成后重建文件分组（使用全新会话）
            if config_id is not None:
                t_rg = time.perf_counter()
                logger.info(f'开始重建文件分组: config_id={config_id}')
                _db_rebuild = db_session_factory()
                try:
                    rebuild_file_groups(config_id, _db_rebuild)
                finally:
                    _db_rebuild.close()
                logger.info(f'重建文件分组完成: config_id={config_id}, 耗时={time.perf_counter() - t_rg:.1f}s')
        except Exception as e:
            logger.exception(f'解析后台任务出错 [{description}]: task_id={task_id}, error={e}')
            try:
                db.rollback()
            except Exception:
                pass
            with parse_tasks_lock:
                parse_tasks[task_id].update({
                    'status': 'error',
                    'message': str(e),
                    'ended_at': datetime.now(),
                })
            # 同步到 UI 日志面板，便于前端直接看到报错
            add_parse_log(task_id, f'任务执行出错: {e}', level='error')
        finally:
            try:
                db.close()
            except Exception:
                pass
            # 清理过期的已完成任务，防止全局字典内存泄漏
            _cleanup_old_tasks()
            logger.info(f'解析后台任务线程结束 [{description}]: task_id={task_id}, '
                        f'总耗时={time.perf_counter() - task_start:.1f}s')

    thread = threading.Thread(target=_background_run, daemon=True)
    thread.start()
    logger.info(f'解析后台任务已启动 [{description}]: task_id={task_id}')

    return task_id


@app.post('/api/parse-cancel/{task_id}')
def cancel_parse_task(task_id: str):
    """取消正在运行的AI解析任务"""
    with parse_tasks_lock:
        task = parse_tasks.get(task_id)
    if task:
        logger.info(f'取消AI解析任务: task_id={task_id}, description="{task.get("description", "")}", status={task.get("status")}')
    else:
        logger.info(f'取消AI解析任务: task_id={task_id} (任务不存在或已完成)')

    with parse_cancel_lock:
        event = parse_cancel_events.get(task_id)
        if event:
            event.set()
    with parse_tasks_lock:
        task = parse_tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail='任务不存在')
    if task['status'] != 'running':
        raise HTTPException(status_code=400, detail='任务不在运行状态')
    return {'success': True, 'message': '已发送取消信号'}


# ===================== AI解析 API =====================


# ===================== AI解析 API =====================


@app.post('/api/parse-name-regex')
def run_parse_name_regex(
    file_ids: List[int] = Query(...),
    concurrency: int = Query(8, ge=1, le=32),
    db: Session = Depends(get_db),
):
    """启动工程方法（正则）文件名解析（后台异步，多进程），不调用AI"""
    if not file_ids:
        raise HTTPException(status_code=400, detail='请选择要解析的文件')

    # 从 file_ids 中获取 config_id（仅用于解析完成后重建分组）
    config_id_row = db.query(ScanResult.scan_config_id).filter(
        ScanResult.id.in_(file_ids)
    ).first()
    parse_config_id = config_id_row[0] if config_id_row else None

    task_id = _start_parse_task(
        parse_file_names_regex_only, f'工程文件名解析 ({len(file_ids)}个)',
        SessionLocal, file_ids, concurrency=concurrency,
        config_id=parse_config_id,
        pymysql_factory=(None if IS_SQLITE else make_pymysql_conn),
    )
    logger.info(f'启动工程文件名解析: task_id={task_id}, file_ids_count={len(file_ids)}, concurrency={concurrency}')
    return {'task_id': task_id, 'message': f'已启动工程文件名解析，共{len(file_ids)}个文件'}


@app.post('/api/parse-name-regex-all/{config_id}')
def run_parse_name_regex_all(
    config_id: int,
    concurrency: int = Query(8, ge=1, le=32),
    force: bool = Query(False),
    db: Session = Depends(get_db),
):
    """启动全部文件的工程方法（正则）文件名解析（多进程 + config_id 分页流式）"""
    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail='扫描配置不存在')

    total = db.query(func.count(ScanResult.id)).filter(
        ScanResult.scan_config_id == config_id
    ).scalar() or 0
    if total == 0:
        raise HTTPException(status_code=400, detail='该配置下没有扫描结果')

    # 「解析全部」不再一次性加载全部 file_ids，交由解析函数按 config_id 分页处理，
    # 避免超大 IN 子句预查询阻塞（见评审报告问题 #2）。
    task_id = _start_parse_task(
        parse_file_names_regex_only, f'工程文件名解析全部 ({total}个)',
        SessionLocal, concurrency=concurrency, force=force,
        config_id=config_id, forward_config_id=True,
        pymysql_factory=(None if IS_SQLITE else make_pymysql_conn),
    )
    logger.info(f'启动全部工程文件名解析: config_id={config_id}, task_id={task_id}, file_count={total}, concurrency={concurrency}, force={force}')
    return {'task_id': task_id, 'message': f'已启动全部工程文件名解析，共{total}个文件'}


@app.post('/api/parse-summary-regex')
def run_parse_summary_regex(
    file_ids: List[int] = Query(...),
    concurrency: int = Query(8, ge=1, le=32),
    db: Session = Depends(get_db),
):
    """启动工程方法（正则）摘要提取（后台异步，多进程），不调用AI"""
    if not file_ids:
        raise HTTPException(status_code=400, detail='请选择要解析的文件')

    # 从 file_ids 中获取 config_id（仅用于解析完成后重建分组）
    config_id_row = db.query(ScanResult.scan_config_id).filter(
        ScanResult.id.in_(file_ids)
    ).first()
    parse_config_id = config_id_row[0] if config_id_row else None

    task_id = _start_parse_task(
        parse_file_summary_regex_only, f'工程摘要解析 ({len(file_ids)}个)',
        SessionLocal, file_ids, concurrency=concurrency,
        config_id=parse_config_id,
        pymysql_factory=(None if IS_SQLITE else make_pymysql_conn),
    )
    logger.info(f'启动工程摘要解析: task_id={task_id}, file_ids_count={len(file_ids)}, concurrency={concurrency}')
    return {'task_id': task_id, 'message': f'已启动工程摘要解析，共{len(file_ids)}个文件'}


@app.post('/api/parse-summary-regex-all/{config_id}')
def run_parse_summary_regex_all(
    config_id: int,
    concurrency: int = Query(8, ge=1, le=32),
    force: bool = Query(False),
    db: Session = Depends(get_db),
):
    """启动全部文件的工程方法（正则）摘要提取（多进程 + config_id 分页流式）"""
    config = db.query(ScanConfig).filter(ScanConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail='扫描配置不存在')

    total = db.query(func.count(ScanResult.id)).filter(
        ScanResult.scan_config_id == config_id
    ).scalar() or 0
    if total == 0:
        raise HTTPException(status_code=400, detail='该配置下没有扫描结果')

    # 「解析全部」不再一次性加载全部 file_ids，交由解析函数按 config_id 分页处理。
    task_id = _start_parse_task(
        parse_file_summary_regex_only, f'工程摘要解析全部 ({total}个)',
        SessionLocal, concurrency=concurrency, force=force,
        config_id=config_id, forward_config_id=True,
        pymysql_factory=(None if IS_SQLITE else make_pymysql_conn),
    )
    logger.info(f'启动全部工程摘要解析: config_id={config_id}, task_id={task_id}, file_count={total}, concurrency={concurrency}, force={force}')
    return {'task_id': task_id, 'message': f'已启动全部工程摘要解析，共{total}个文件'}


@app.get('/api/parse-tasks/running')
def get_running_parse_task():
    """获取当前正在运行的解析任务（没有则返回null）"""
    with parse_tasks_lock:
        for tid, task in parse_tasks.items():
            if task['status'] == 'running':
                result = {
                    'task_id': tid,
                    'status': task['status'],
                    'total': task['total'],
                    'processed': task['processed'],
                    'success': task['success'],
                    'failed': task['failed'],
                    'description': task['description'],
                }
                if task.get('started_at'):
                    elapsed = (datetime.now() - task['started_at']).total_seconds()
                    result['elapsed'] = elapsed
                with parse_logs_lock:
                    logs = parse_logs.get(tid, [])[-100:]
                result['logs'] = logs
                return result
    return {'task_id': None, 'status': None}


@app.get('/api/parse-tasks/{task_id}')
def get_parse_task_progress(task_id: str, db: Session = Depends(get_db)):
    """获取AI解析任务进度"""
    with parse_tasks_lock:
        task = parse_tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail='任务不存在')

    result = {
        'status': task['status'],
        'total': task['total'],
        'processed': task['processed'],
        'success': task['success'],
        'failed': task['failed'],
        'description': task['description'],
    }

    # 计算已用时间
    if task.get('started_at'):
        elapsed = (datetime.now() - task['started_at']).total_seconds()
        result['elapsed'] = elapsed

    if task['status'] == 'done':
        result['result'] = task['result']
    elif task['status'] == 'error':
        result['message'] = task.get('message', '未知错误')

    # 优先从内存获取日志（运行中任务），内存没有则从数据库读取（已完成/过期任务）
    with parse_logs_lock:
        logs = parse_logs.get(task_id, [])
    if not logs:
        try:
            db_logs = db.query(ParseLog).filter(
                ParseLog.task_id == task_id
            ).order_by(ParseLog.id.asc()).limit(100).all()
            logs = [{'time': l.created_at.strftime('%H:%M:%S') if l.created_at else '',
                     'msg': l.message, 'level': l.level} for l in db_logs]
        except Exception:
            logs = []
    else:
        logs = logs[-100:]

    result['logs'] = logs

    return result


# ===================== 导出 Markdown API =====================

@app.post('/api/export-md')
def export_markdown(
    ids: List[int] = Query(default=None),
    config_id: int = Query(default=None),
    db: Session = Depends(get_db),
):
    """导出选中文件的解析结果为Markdown文件。
    支持两种方式：① ids=1,2,3 导出指定记录；② config_id=N 导出整个配置的记录。
    """
    if not ids and not config_id:
        raise HTTPException(status_code=400, detail='请提供 ids 或 config_id')

    query = db.query(
        ScanResult.id,
        ScanResult.file_name,
        FileMetadata.novel_name,
        FileMetadata.author,
        FileMetadata.summary,
        FileMetadata.progress,
        FileMetadata.source,
    ).outerjoin(
        FileMetadata, ScanResult.id == FileMetadata.scan_result_id
    )
    if ids:
        query = query.filter(ScanResult.id.in_(ids))
    elif config_id:
        query = query.filter(ScanResult.scan_config_id == config_id)
    records = query.all()

    if not records:
        raise HTTPException(status_code=404, detail='未找到记录')

    export_dir = os.path.join(app_data_root(), 'exports')
    os.makedirs(export_dir, exist_ok=True)

    exported_files = []

    name_counts = {}
    for r in records:
        # 优先使用解析后的小说名（novel_name），否则用源文件名。
        # 关键修复：r.novel_name 来自 DB 解析结果，可能含 Windows 非法字符
        # （? / \ : * 等），旧代码直接作为文件名会导致 open() 抛 OSError。
        # 现在统一走 _sanitize_md_filename 清洗。
        raw_name = (r.novel_name or '').strip()
        if not raw_name:
            raw_name = os.path.splitext(r.file_name)[0]
        safe_name = _sanitize_md_filename(raw_name)
        if not safe_name:
            # 极端兜底：原始名清洗后全空，用 id 区分避免互相覆盖
            safe_name = f"未命名_{r.id}"

        # 同名文件去重
        if safe_name in name_counts:
            name_counts[safe_name] += 1
            safe_name = f"{safe_name}_{name_counts[safe_name]}"
        else:
            name_counts[safe_name] = 0

        md_content = _build_md_export(r)
        md_path = os.path.join(export_dir, f"{safe_name}.md")

        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(md_content)

        exported_files.append({
            'file_name': f"{safe_name}.md",
            'path': md_path,
        })

    logger.info(f'导出Markdown: {len(exported_files)}个文件, 目录={export_dir}')
    return {
        'exported_count': len(exported_files),
        'files': exported_files,
        'directory': export_dir,
        'message': f'成功导出{len(exported_files)}个Markdown文件',
    }


# ===================== 导出 Excel API =====================

@app.post('/api/export-excel')
def export_excel(ids: Optional[List[int]] = Query(None),
                 config_id: Optional[int] = Query(None),
                 db: Session = Depends(get_db)):
    """导出解析结果为Excel文件
    支持按ids列表导出选中记录，或按config_id导出全部记录
    """
    query = db.query(
        ScanResult.id,
        ScanResult.file_name,
        ScanResult.file_size,
        ScanResult.file_path,
        FileMetadata.novel_name,
        FileMetadata.author,
        FileMetadata.summary,
        FileMetadata.progress,
        FileMetadata.source,
    ).outerjoin(
        FileMetadata, ScanResult.id == FileMetadata.scan_result_id
    )

    if ids:
        query = query.filter(ScanResult.id.in_(ids))
    elif config_id:
        query = query.filter(ScanResult.scan_config_id == config_id)
    else:
        raise HTTPException(status_code=400, detail='请指定要导出的文件（ids或config_id）')

    records = query.all()

    if not records:
        raise HTTPException(status_code=404, detail='未找到记录')

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail='导出Excel需要安装openpyxl: pip install openpyxl')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '解析结果'

    # 表头
    headers = ['ID', '文件名', '小说名', '作者', '进度', '来源', '摘要']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 数据行
    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.file_name)
        ws.cell(row=row_idx, column=3, value=r.novel_name or '')
        ws.cell(row=row_idx, column=4, value=r.author or '')
        ws.cell(row=row_idx, column=5, value=getattr(r, 'progress', '') or '')
        ws.cell(row=row_idx, column=6, value=getattr(r, 'source', '') or '')
        ws.cell(row=row_idx, column=7, value=r.summary or '')

    # 列宽
    col_widths = [8, 50, 25, 15, 18, 12, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 自适应行高
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and len(cell.value) > 50:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    export_dir = os.path.join(app_data_root(), 'exports')
    os.makedirs(export_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = os.path.join(export_dir, f'解析结果_{timestamp}.xlsx')
    wb.save(file_path)

    logger.info(f'导出Excel: {len(records)}条记录, {file_path}')
    return {
        'exported_count': len(records),
        'file_name': f'解析结果_{timestamp}.xlsx',
        'directory': export_dir,
        'message': f'成功导出{len(records)}条记录到Excel',
    }


# ===================== 帮助手册 API =====================

@app.get('/api/help')
def list_help_docs(db: Session = Depends(get_db)):
    """获取帮助手册文档列表（不含正文）"""
    docs = db.query(HelpDoc).order_by(HelpDoc.sort_order, HelpDoc.id).all()
    return [
        {
            'id': d.id,
            'doc_key': d.doc_key,
            'title': d.title,
            'sort_order': d.sort_order,
            'updated_at': d.updated_at.isoformat() if d.updated_at else None,
        }
        for d in docs
    ]


@app.get('/api/help/{doc_key}')
def get_help_doc(doc_key: str, db: Session = Depends(get_db)):
    """获取单篇帮助文档（含 Markdown 正文）"""
    doc = db.query(HelpDoc).filter(HelpDoc.doc_key == doc_key).first()
    if not doc:
        raise HTTPException(status_code=404, detail='文档不存在')
    return {
        'id': doc.id,
        'doc_key': doc.doc_key,
        'title': doc.title,
        'content': doc.content,
        'sort_order': doc.sort_order,
        'updated_at': doc.updated_at.isoformat() if doc.updated_at else None,
    }


# ===================== 列配置 API =====================

@app.get('/api/column-configs')
def list_column_configs(db: Session = Depends(get_db)):
    """获取列显示配置"""
    cols = db.query(ColumnConfig).order_by(ColumnConfig.sort_order).all()
    return [
        {
            'id': c.id,
            'column_key': c.column_key,
            'display_name': c.display_name,
            'visible': c.visible,
            'sort_order': c.sort_order,
        }
        for c in cols
    ]


@app.put('/api/column-configs')
def update_column_configs(data: List[dict], db: Session = Depends(get_db)):
    """批量更新列显示配置"""
    for item in data:
        col = db.query(ColumnConfig).filter(ColumnConfig.column_key == item['column_key']).first()
        if col:
            if 'visible' in item:
                col.visible = item['visible']
            if 'sort_order' in item:
                col.sort_order = item['sort_order']
    db.commit()
    logger.info('列显示配置已更新')
    return {'message': '更新成功'}


# ===================== 系统信息 API =====================

# ===================== 关键词替换规则 API =====================
class KeywordReplaceRuleIn(BaseModel):
    scope: str                       # 'scan' | 'parse'
    pattern: str
    replacement: str = ''
    sort_order: int = 0
    enabled: bool = True


@app.get('/api/keyword-replaces')
def list_keyword_rules(db: Session = Depends(get_db)):
    """列出全部关键词替换规则（设置页展示，按作用域、顺序排序）"""
    rules = db.query(KeywordReplaceRule).order_by(
        KeywordReplaceRule.scope, KeywordReplaceRule.sort_order, KeywordReplaceRule.id
    ).all()
    return {'rules': [{
        'id': r.id,
        'scope': r.scope,
        'pattern': r.pattern,
        'replacement': r.replacement,
        'sort_order': r.sort_order,
        'enabled': r.enabled,
    } for r in rules]}


@app.post('/api/keyword-replaces')
def create_keyword_rule(rule: KeywordReplaceRuleIn, db: Session = Depends(get_db)):
    scope = rule.scope if rule.scope in ('scan', 'parse') else 'scan'
    obj = KeywordReplaceRule(
        scope=scope,
        pattern=rule.pattern,
        replacement=rule.replacement or '',
        sort_order=rule.sort_order,
        enabled=rule.enabled,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    logger.info(f'新增关键词替换规则: scope={scope}, pattern={rule.pattern!r}')
    return {'id': obj.id, 'message': '已添加规则'}


@app.put('/api/keyword-replaces/{rule_id}')
def update_keyword_rule(rule_id: int, rule: KeywordReplaceRuleIn, db: Session = Depends(get_db)):
    obj = db.query(KeywordReplaceRule).filter(KeywordReplaceRule.id == rule_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail='规则不存在')
    obj.scope = rule.scope if rule.scope in ('scan', 'parse') else obj.scope
    obj.pattern = rule.pattern
    obj.replacement = rule.replacement or ''
    obj.sort_order = rule.sort_order
    obj.enabled = rule.enabled
    db.commit()
    logger.info(f'更新关键词替换规则: id={rule_id}, pattern={rule.pattern!r}')
    return {'message': '已更新规则'}


@app.delete('/api/keyword-replaces/{rule_id}')
def delete_keyword_rule(rule_id: int, db: Session = Depends(get_db)):
    obj = db.query(KeywordReplaceRule).filter(KeywordReplaceRule.id == rule_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail='规则不存在')
    db.delete(obj)
    db.commit()
    logger.info(f'删除关键词替换规则: id={rule_id}')
    return {'message': '已删除规则'}


# ===================== 一键清理流程 API =====================
class PipelineStartIn(BaseModel):
    config_id: int
    delete_mode: str = 'db'     # 'db'=仅删数据库记录；'file'=数据库+源文件


@app.post('/api/pipeline/start')
def api_pipeline_start(data: PipelineStartIn):
    """启动一键清理（单实例）。先选删除类型，再跑 扫描→解析→合集→标记重复→清理"""
    delete_mode = data.delete_mode if data.delete_mode in ('db', 'file') else 'db'
    try:
        tid = pipeline_manager.start_pipeline(data.config_id, delete_mode)
    except RuntimeError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return {'task_id': tid, 'message': '已启动一键清理'}


@app.get('/api/pipeline/status')
def api_pipeline_status():
    """查询当前流程各节点状态/进度/时间"""
    return pipeline_manager.get_status()


@app.post('/api/pipeline/cancel')
def api_pipeline_cancel():
    """取消进行中的流程（已完成节点保留，不回滚）"""
    ok = pipeline_manager.cancel_pipeline()
    return {'success': ok}


@app.post('/api/pipeline/confirm')
def api_pipeline_confirm():
    """二次确认：确认后真正执行删除"""
    ok = pipeline_manager.confirm_pipeline()
    if not ok:
        raise HTTPException(status_code=400, detail='当前没有待确认删除的流程')
    return {'success': True}


@app.get('/api/pipeline/logs')
def api_pipeline_logs():
    """回溯查询流程日志（持久化于 ParseLog 表）"""
    return {'logs': pipeline_manager.get_logs()}


@app.get('/api/stats')
def get_stats(db: Session = Depends(get_db)):
    """获取统计信息"""
    total_files = db.query(ScanResult).count()
    total_configs = db.query(ScanConfig).count()
    parsed_count = db.query(FileMetadata).filter(
        FileMetadata.novel_name != ''
    ).count()
    return {
        'total_files': total_files,
        'total_configs': total_configs,
        'parsed_count': parsed_count,
    }


# ===================== 前端静态文件 =====================

@app.get('/')
def serve_index():
    index_path = os.path.join(FRONTEND_DIR, 'index.html')
    if os.path.isfile(index_path):
        return FileResponse(index_path)
    return {'status': 'running'}


# 挂载前端静态文件（CSS/JS等）
if os.path.isdir(FRONTEND_DIR):
    app.mount('/static', StaticFiles(directory=FRONTEND_DIR), name='static')

# 挂载导出文件目录（Excel/MD 导出产物），供前端 window.open('/exports/xxx') 下载
_exports_dir = os.path.join(app_data_root(), 'exports')
os.makedirs(_exports_dir, exist_ok=True)
app.mount('/exports', StaticFiles(directory=_exports_dir), name='exports')


# ===================== 工具函数 =====================


def _sanitize_md_filename(name, max_length: int = 80) -> str:
    r"""清洗用于 Markdown 导出的文件名（处理 Windows 非法字符 + 边界情况）。

    策略（黑名单 + 保留其它一切 Unicode 字符，包括中英文标点）：
      - 用黑名单替换 Windows 保留的非法字符：< > : " / \ | ? *
      - 替换控制字符（ord < 32，如 \\r \\n \\t）
      - 去除首尾的空白、点、下划线（Windows 文件名规则）
      - 限制长度避免超长路径
      - 空结果回退为"未命名"避免覆盖或写入失败

    与早期实现的区别：旧版用"白名单 isalnum + 极少标点"导致 novel_name 含
    《》【】、「」等中文书名号/方括号时被错误替换成下划线；
    本版采用黑名单方式，可保留常见中文标点，又避免含 Windows 非法字符时写入失败。
    """
    if not name:
        return ''
    # Windows 保留的非法字符集（注意：全角冒号"：" / 全角问号"？"
    # 在 Windows 文件名中是合法的，常见于小说标题，不应被替换）
    illegal = '<>:"/\\|?*'
    cleaned_chars = []
    for c in str(name):
        if c in illegal or ord(c) < 32:
            cleaned_chars.append('_')
        else:
            cleaned_chars.append(c)
    result = ''.join(cleaned_chars).strip(' ._')
    if len(result) > max_length:
        result = result[:max_length].rstrip(' ._')
    return result or '未命名'


def _build_md_export(r) -> str:
    """构建单个文件的Markdown导出内容"""
    lines = []
    lines.append(f'# {r.novel_name or r.file_name}')
    lines.append('')

    if r.author:
        lines.append(f'**作者**: {r.author}')
        lines.append('')
    if getattr(r, 'progress', ''):
        lines.append(f'**进度**: {r.progress}')
        lines.append('')
    if getattr(r, 'source', ''):
        lines.append(f'**来源**: {r.source}')
        lines.append('')
    if r.summary:
        lines.append('---')
        lines.append('## 内容简介')
        lines.append('')
        lines.append(r.summary)
        lines.append('')

    return '\n'.join(lines)


def _build_result_item(r) -> dict:
    """构建单条结果项，返回给前端"""
    created_date_str = ''
    if r.created_date:
        created_date_str = r.created_date.strftime('%Y-%m-%d %H:%M')

    return {
        'id': r.id,
        'file_name': r.file_name or '',
        'file_size': r.file_size or 0,
        'file_size_display': _format_size(r.file_size or 0),
        'file_path': r.file_path or '',
        'created_date': created_date_str,
        'novel_name': r.novel_name or '',
        'author': r.author or '',
        'summary': r.summary or '',
        'progress': getattr(r, 'progress', '') or '',
        'source': getattr(r, 'source', '') or '',
        'scan_config_id': r.scan_config_id,
    }


def _format_size(size: int) -> str:
    """格式化文件大小显示"""
    if size < 1024:
        return f'{size} B'
    elif size < 1024 ** 2:
        return f'{size / 1024:.1f} KB'
    elif size < 1024 ** 3:
        return f'{size / 1024 ** 2:.1f} MB'
    else:
        return f'{size / 1024 ** 3:.2f} GB'


# ===================== 启动入口 =====================
if __name__ == '__main__':
    import uvicorn
    logger.info('启动服务: http://0.0.0.0:8000')
    uvicorn.run(app, host='0.0.0.0', port=8000, log_level='info')
